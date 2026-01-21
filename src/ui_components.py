"""
UI Components Module

Contains all Streamlit UI rendering functions for step-by-step processing.
Helper functions and sidebar are in the ui/ package.
"""

import os
import re
import json
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

import streamlit as st

from state_management import (
    SOURCE_DIR, get_pdf_slug, get_output_dir, get_images_dir, get_source_dir,
    get_available_textbooks, clear_session_data, load_saved_data,
    load_settings, load_qc_progress, save_settings, save_chapters,
    save_questions, save_images, save_pages, save_image_assignments,
    save_qc_progress, get_raw_blocks_file, save_raw_blocks, save_question_blocks,
    save_global_settings
)
from pdf_extraction import (
    extract_images_from_pdf, assign_chapters_to_images,
    extract_chapter_text, render_pdf_page,
    extract_text_with_lines, insert_image_markers, build_chapter_text_with_lines,
    extract_lines_by_range_mapped, get_pages_for_line_range
)
from llm_extraction import (
    get_anthropic_client, get_model_options,
    identify_chapters_llm, add_page_numbers_to_questions,
    load_prompts, save_prompts, reload_prompts,
    get_extraction_logger, reset_logger,
    identify_question_blocks_llm, format_raw_block_llm
)
from cost_tracking import (
    get_session_summary, get_step_cost, format_cost_display,
    format_cost, format_tokens, save_cost_tracking,
    set_last_step_cost, get_last_step_cost, clear_last_step_cost,
    flush_pending_costs
)

# Import from modular ui package
from ui.helpers import (
    play_completion_sound,
    build_block_aware_image_assignments,
    get_selected_model_id,
    clear_step_data,
    sort_chapter_keys,
    question_sort_key,
    get_images_for_question,
    get_answer_images_for_question,
    get_all_question_options,
)
from ui.sidebar import render_sidebar

# Note: Helper functions and sidebar are now in ui/helpers.py and ui/sidebar.py
# They are imported above from the ui package


# =============================================================================
# Cost Display Helper
# =============================================================================

def render_cost_metrics(step_name: str = None, show_session: bool = False):
    """
    Store cost metrics for display after rerun.

    Args:
        step_name: If provided, store cost for specific step.
        show_session: If True, also show session totals alongside step cost.
    """
    # First flush any pending costs from worker threads
    pending = flush_pending_costs()

    if step_name:
        step_data = get_step_cost(step_name)
        input_tokens = step_data["input"]
        output_tokens = step_data["output"]
        cost = step_data["cost"]

        # If no step data but we have pending costs, use those
        if cost == 0 and pending["cost"] > 0:
            input_tokens = pending["input_tokens"]
            output_tokens = pending["output_tokens"]
            cost = pending["cost"]

        if cost > 0:
            # Store for display after rerun
            set_last_step_cost(step_name, input_tokens, output_tokens, cost)


def display_last_step_cost():
    """Display the cost from the last completed step, if available."""
    last_cost = get_last_step_cost()
    if last_cost:
        step = last_cost["step"]
        input_tokens = last_cost["input_tokens"]
        output_tokens = last_cost["output_tokens"]
        cost = last_cost["cost"]
        st.info(f"**Last step cost ({step}):** {format_tokens(input_tokens)} in / {format_tokens(output_tokens)} out | {format_cost(cost)}")
        # Clear after displaying so it doesn't persist
        clear_last_step_cost()


# =============================================================================
# Step 1: Source Selection
# =============================================================================

def render_source_step():
    """Render source PDF selection step."""
    col1, col2 = st.columns([4, 1])
    with col1:
        st.header("Step 1: Select Source PDF")
    with col2:
        if st.button("Reset To This Step", key="clear_source", type="secondary", help="Clear this step and all subsequent steps"):
            clear_step_data("source")
            st.success("All data cleared")
            st.rerun()

    # Folder selection section
    st.subheader("Source Folder")

    current_source = st.session_state.get("source_dir", SOURCE_DIR)

    col_path, col_btn = st.columns([4, 1])
    with col_path:
        st.text_input("Current folder:", value=current_source, disabled=True, key="source_folder_display")
    with col_btn:
        if st.button("Browse...", key="browse_folder_btn"):
            # Open native macOS Finder folder picker via AppleScript
            import subprocess

            # Set initial directory for the dialog
            initial_dir = current_source if os.path.exists(current_source) else str(Path.home())

            script = f'''
            set defaultFolder to POSIX file "{initial_dir}" as alias
            try
                set selectedFolder to choose folder with prompt "Select Source Folder" default location defaultFolder
                return POSIX path of selectedFolder
            on error
                return ""
            end try
            '''

            result = subprocess.run(
                ['osascript', '-e', script],
                capture_output=True,
                text=True
            )

            selected = result.stdout.strip()
            if selected:
                st.session_state.source_dir = selected.rstrip('/')
                save_global_settings()
                st.rerun()

    st.markdown("---")

    # Use selected source directory for PDF scanning
    source_dir = get_source_dir()
    pdf_files = list(Path(source_dir).glob("*.pdf")) if os.path.exists(source_dir) else []

    if not pdf_files:
        st.warning(f"No PDF files found in '{source_dir}'. Please select a folder containing PDF files.")
        return

    pdf_options = sorted([f.name for f in pdf_files])
    available_textbooks = get_available_textbooks()

    # Default to last selected PDF if available
    last_pdf = st.session_state.get("last_selected_pdf", "")
    default_idx = 0
    if last_pdf in pdf_options:
        default_idx = pdf_options.index(last_pdf)

    selected_pdf = st.selectbox("Select PDF file:", pdf_options, index=default_idx)

    if selected_pdf:
        pdf_path = f"{source_dir}/{selected_pdf}"
        output_slug = get_pdf_slug(selected_pdf)
        st.info(f"Selected: {pdf_path}")
        st.caption(f"Output folder: output/{output_slug}/")

        has_existing_data = output_slug in available_textbooks

        col1, col2 = st.columns(2)

        with col1:
            btn_label = "Load PDF (Fresh Start)" if has_existing_data else "Load PDF"
            if st.button(btn_label, type="primary"):
                st.session_state.current_pdf = selected_pdf
                st.session_state.last_selected_pdf = selected_pdf
                save_global_settings()
                clear_session_data()
                reset_logger()  # Reset logger for new PDF

                with st.spinner("Extracting images from PDF..."):
                    st.session_state.images = extract_images_from_pdf(pdf_path, get_images_dir())
                    save_images()

                with st.spinner("Extracting text with image markers..."):
                    # Extract text with positions for accurate image marker placement
                    pages, all_lines = extract_text_with_lines(pdf_path)
                    st.session_state.pages = pages
                    st.session_state.pdf_path = pdf_path
                    save_pages()

                    # Insert image markers at correct positions
                    lines_with_markers = insert_image_markers(all_lines, st.session_state.images, pages)

                    # Build the full document text with line numbers
                    numbered_lines = []
                    line_num = 1
                    for line in lines_with_markers:
                        if line.startswith("[IMAGE:"):
                            numbered_lines.append(line)
                        else:
                            numbered_lines.append(f"[LINE:{line_num:06d}] {line}")
                            line_num += 1

                    full_text = "\n".join(numbered_lines)
                    st.session_state.extracted_text = full_text

                    # Save to file
                    output_dir = get_output_dir()
                    text_file_path = os.path.join(output_dir, "extracted_text.txt")
                    with open(text_file_path, "w", encoding="utf-8") as f:
                        f.write(full_text)

                st.success(f"Loaded {len(st.session_state.pages)} pages, {len(st.session_state.images)} images")
                st.info(f"Saved extracted text to: {text_file_path}")
                st.rerun()

        with col2:
            if has_existing_data:
                if st.button("Load Existing Progress"):
                    st.session_state.current_pdf = selected_pdf
                    st.session_state.last_selected_pdf = selected_pdf
                    save_global_settings()
                    clear_session_data()
                    reset_logger()  # Reset logger for loaded textbook
                    load_saved_data()
                    load_settings()
                    st.session_state.qc_progress = load_qc_progress()

                    # Set pdf_path from current source folder and selected PDF
                    st.session_state.pdf_path = pdf_path

                    # Load extracted text from file if it exists
                    text_file_path = os.path.join(get_output_dir(), "extracted_text.txt")
                    if os.path.exists(text_file_path):
                        with open(text_file_path, "r", encoding="utf-8") as f:
                            st.session_state.extracted_text = f.read()

                    st.success("Loaded previous session data")
                    st.rerun()

        if st.session_state.pages:
            st.success(f"PDF loaded: {len(st.session_state.pages)} pages, {len(st.session_state.images)} images")
            st.markdown("**Next:** Go to 'Extract Chapters' to identify chapter boundaries.")

            # Full extracted text with image markers
            st.markdown("---")
            st.subheader("Extracted Text with Image Markers")

            extracted_text = st.session_state.get("extracted_text")
            if extracted_text:
                # Show stats
                line_count = extracted_text.count("[LINE:")
                image_count = extracted_text.count("[IMAGE:")
                st.caption(f"{line_count} text lines, {image_count} image markers")

                # Show file path
                text_file_path = os.path.join(get_output_dir(), "extracted_text.txt")
                st.caption(f"File: `{text_file_path}`")

                # Full scrollable text area
                st.text_area(
                    "Full document (scroll to view):",
                    extracted_text,
                    height=600,
                    key="full_extracted_text"
                )
            else:
                st.warning("Extracted text not available. Try re-loading the PDF.")


# =============================================================================
# Step 2: Extract Chapters
# =============================================================================

def render_chapters_step():
    """Render chapter extraction step."""
    col1, col2 = st.columns([4, 1])
    with col1:
        st.header("Step 2: Extract Chapters")
    with col2:
        if st.button("Reset To This Step", key="clear_chapters", type="secondary", help="Clear this step and all subsequent steps"):
            clear_step_data("chapters")
            st.success("Chapters and subsequent data cleared")
            st.rerun()

    # Show cost from last completed step
    display_last_step_cost()

    has_pages = st.session_state.pages is not None
    has_chapters = st.session_state.chapters is not None

    if not has_pages and not has_chapters:
        st.warning("Please load a PDF first (Step 1)")
        return

    client = get_anthropic_client()
    if not client:
        st.error("ANTHROPIC_API_KEY not set. Please configure your .env file.")
        return

    if has_pages:
        col1, col2 = st.columns(2)

        with col1:
            model_options = get_model_options()
            current_idx = model_options.index(st.session_state.selected_model) if st.session_state.selected_model in model_options else 0
            selected_model = st.selectbox("Model:", model_options, index=current_idx, key="chapters_model")
            if selected_model != st.session_state.selected_model:
                st.session_state.selected_model = selected_model
                save_settings()

        with col2:
            btn_label = "Re-extract Chapters" if has_chapters else "Extract Chapters"
            if st.button(btn_label, type="primary"):
                with st.spinner(f"Using {st.session_state.selected_model} to identify chapters..."):
                    chapters = identify_chapters_llm(client, st.session_state.pages, get_selected_model_id())
                    # Sort chapters by chapter_number to ensure correct ordering
                    chapters = sorted(chapters, key=lambda ch: ch.get("chapter_number", 0))
                    st.session_state.chapters = chapters

                    for i, ch in enumerate(chapters):
                        start_page = ch["start_page"]
                        end_page = chapters[i + 1]["start_page"] if i + 1 < len(chapters) else None
                        ch_key = f"ch{ch['chapter_number']}"
                        st.session_state.chapter_texts[ch_key] = extract_chapter_text(
                            st.session_state.pages, start_page, end_page
                        )

                    save_chapters()

                    if st.session_state.images:
                        st.session_state.images = assign_chapters_to_images(
                            st.session_state.images, chapters
                        )
                        save_images()

                    # Save cost tracking data
                    save_cost_tracking(get_output_dir())

                st.success(f"Found {len(chapters)} chapters")
                render_cost_metrics("identify_chapters")
                play_completion_sound()
                st.rerun()
    elif has_chapters:
        st.info("Chapters already extracted. Go to Step 1 to reload PDF if you need to re-extract.")

    if st.session_state.chapters:
        st.subheader("Extracted Chapters")

        for ch in st.session_state.chapters:
            st.markdown(f"**Chapter {ch['chapter_number']}:** {ch['title']} (page {ch['start_page']})")

        st.markdown("---")

        st.subheader("Preview Chapter Text")

        chapter_options = [f"Ch{ch['chapter_number']}: {ch['title'][:40]}..."
                         for ch in st.session_state.chapters]
        selected_ch_idx = st.selectbox("Select chapter to preview:",
                                        range(len(chapter_options)),
                                        format_func=lambda x: chapter_options[x])

        if selected_ch_idx is not None:
            ch = st.session_state.chapters[selected_ch_idx]
            ch_key = f"ch{ch['chapter_number']}"
            ch_text = st.session_state.chapter_texts.get(ch_key, "")

            st.text_area("Chapter text:", ch_text, height=400, disabled=True)
            st.caption(f"Text length: {len(ch_text):,} characters")


# =============================================================================
# Step 3: Extract Questions (Raw)
# =============================================================================

def render_questions_step():
    """Render raw question extraction step - identifies line ranges and extracts raw text."""
    col1, col2 = st.columns([4, 1])
    with col1:
        st.header("Step 3: Extract Questions")
    with col2:
        if st.button("Reset To This Step", key="clear_questions", type="secondary", help="Clear this step and all subsequent steps"):
            clear_step_data("questions")
            st.success("Questions and subsequent data cleared")
            st.rerun()

    # Show cost from last completed step
    display_last_step_cost()

    st.markdown("""
    This step identifies question and answer boundaries in the text and extracts raw Q&A pairs.
    The LLM identifies line ranges for each question/answer, then the raw text is extracted.
    """)

    # Block extraction is always enabled (v2 mode)
    st.info("**Block extraction mode:** LLM identifies block boundaries, raw text preserved with line numbers for Step 4 formatting.")

    if not st.session_state.chapters:
        st.warning("Please extract chapters first (Step 2)")
        return

    client = get_anthropic_client()
    if not client:
        st.error("ANTHROPIC_API_KEY not set. Please configure your .env file.")
        return

    # Initialize logger
    output_dir = get_output_dir()
    logger = get_extraction_logger(output_dir)

    # Chapter selector
    chapter_options = [f"Ch{ch['chapter_number']}: {ch['title'][:40]}..."
                      for ch in st.session_state.chapters]
    selected_ch_idx = st.selectbox("Select chapter:",
                                    range(len(chapter_options)),
                                    format_func=lambda x: chapter_options[x],
                                    key="extract_ch_selector")

    ch = st.session_state.chapters[selected_ch_idx]
    ch_num = ch["chapter_number"]
    ch_key = f"ch{ch_num}"

    # Model selection, workers, and buttons
    model_col, workers_col, btn_col1, btn_col2 = st.columns([2, 1.5, 2, 2])

    with model_col:
        model_options = get_model_options()
        current_idx = model_options.index(st.session_state.selected_model) if st.session_state.selected_model in model_options else 0
        selected_model = st.selectbox("Model:", model_options, index=current_idx, key="questions_model")
        if selected_model != st.session_state.selected_model:
            st.session_state.selected_model = selected_model
            save_settings()

    with workers_col:
        extract_workers = st.number_input(
            "Parallel workers:",
            min_value=1,
            max_value=50,
            value=20,
            help="Number of chapters to extract in parallel. Tier 1: 5-10 | Tier 2+: 10-20",
            key="extract_workers"
        )

    def extract_blocks_from_chapter(ch_idx: int, pages_with_lines: list, lines_with_images: list, on_progress=None):
        """Extract question blocks with raw text and line numbers preserved."""
        ch = st.session_state.chapters[ch_idx]
        ch_num = ch["chapter_number"]
        ch_key = f"ch{ch_num}"
        model_id = get_selected_model_id()

        start_page = ch["start_page"]
        end_page = st.session_state.chapters[ch_idx + 1]["start_page"] if ch_idx + 1 < len(st.session_state.chapters) else None
        ch_text, line_mapping = build_chapter_text_with_lines(
            lines_with_images, pages_with_lines, start_page, end_page
        )

        # Use the new simpler block identification
        blocks = identify_question_blocks_llm(get_anthropic_client(), ch_num, ch_text, model_id, on_progress=on_progress)

        if not blocks:
            logger.warning(f"Chapter {ch_num}: No blocks extracted")
            return []

        raw_blocks = []
        for block in blocks:
            block_id = block.get("block_id", "?")
            q_start = block.get("question_start", 0)
            q_end = block.get("question_end", 0)
            a_start = block.get("answer_start", 0)
            a_end = block.get("answer_end", 0)

            # Extract raw text WITH line numbers preserved for traceability
            question_text_raw = extract_lines_by_range_mapped(
                lines_with_images, q_start, q_end, line_mapping, preserve_line_numbers=True
            ) if q_start > 0 else ""

            answer_text_raw = extract_lines_by_range_mapped(
                lines_with_images, a_start, a_end, line_mapping, preserve_line_numbers=True
            ) if a_start > 0 else ""

            # Calculate page numbers from line ranges
            q_pages = get_pages_for_line_range(q_start, q_end, pages_with_lines)
            a_pages = get_pages_for_line_range(a_start, a_end, pages_with_lines)

            raw_blocks.append({
                "block_id": f"ch{ch_num}_{block_id}",
                "block_label": block_id,
                "chapter": ch_num,
                "question_start": q_start,
                "question_end": q_end,
                "answer_start": a_start,
                "answer_end": a_end,
                "question_pages": q_pages,
                "answer_pages": a_pages,
                "question_text_raw": question_text_raw,  # Raw text with line numbers
                "answer_text_raw": answer_text_raw,       # Raw text with line numbers
                "formatted": False  # Will be set to True after Step 4 formatting
            })

        return raw_blocks

    with btn_col1:
        if st.button(f"Extract Chapter {ch_num}", type="primary"):
            progress_text = st.empty()
            progress_text.text(f"Extracting Chapter {ch_num}... Preparing text...")

            pdf_path = st.session_state.get("pdf_path")
            if not pdf_path or not os.path.exists(pdf_path):
                st.error("PDF path not found. Please reload the PDF in Step 1.")
            else:
                pages_with_lines, all_lines = extract_text_with_lines(pdf_path)
                lines_with_images = insert_image_markers(
                    all_lines, st.session_state.images, pages_with_lines
                )

                progress_text.text(f"Extracting Chapter {ch_num}... Streaming response...")

                # Block extraction with raw text and line numbers preserved
                raw_blocks = extract_blocks_from_chapter(selected_ch_idx, pages_with_lines, lines_with_images)
                if raw_blocks:
                    st.session_state.raw_blocks[ch_key] = raw_blocks
                    save_raw_blocks()
                    save_cost_tracking(get_output_dir())
                    progress_text.empty()
                    st.success(f"Extracted {len(raw_blocks)} blocks from Chapter {ch_num}")
                    render_cost_metrics("identify_blocks")
                    play_completion_sound()
                else:
                    progress_text.empty()
                    st.warning(f"No blocks extracted from Chapter {ch_num}")
                st.rerun()

    with btn_col2:
        if st.button("Extract ALL Chapters"):
            progress_bar = st.progress(0)
            status_text = st.empty()
            chapter_status = st.empty()

            model_id = get_selected_model_id()
            total_chapters = len(st.session_state.chapters)
            max_workers = min(extract_workers, total_chapters)

            # Prepare text with line numbers and image markers
            status_text.text("Preparing text with line numbers and image markers...")
            pdf_path = st.session_state.get("pdf_path")

            if not pdf_path or not os.path.exists(pdf_path):
                st.error("PDF path not found. Please reload the PDF in Step 1.")
                return

            pages_with_lines, all_lines = extract_text_with_lines(pdf_path)
            lines_with_images = insert_image_markers(
                all_lines, st.session_state.images, pages_with_lines
            )
            logger.info(f"Prepared {len(lines_with_images)} lines with {len(st.session_state.images)} image markers")

            # Precompute chapter data for all chapters
            status_text.text("Preparing chapter data...")
            chapter_data = []
            for i, ch in enumerate(st.session_state.chapters):
                ch_num = ch["chapter_number"]
                start_page = ch["start_page"]
                end_page = st.session_state.chapters[i + 1]["start_page"] if i + 1 < len(st.session_state.chapters) else None
                ch_text, line_mapping = build_chapter_text_with_lines(
                    lines_with_images, pages_with_lines, start_page, end_page
                )
                chapter_data.append({
                    "ch_num": ch_num,
                    "ch_key": f"ch{ch_num}",
                    "ch_text": ch_text,
                    "line_mapping": line_mapping
                })

            # Worker function for block extraction with raw text and line numbers
            def extract_block_worker(ch_data: dict) -> tuple[str, list]:
                """Extract blocks from a chapter. Returns (ch_key, raw_blocks)."""
                ch_num = ch_data["ch_num"]
                ch_key = ch_data["ch_key"]
                ch_text = ch_data["ch_text"]
                line_mapping = ch_data["line_mapping"]

                # Use block identification LLM
                blocks = identify_question_blocks_llm(
                    get_anthropic_client(), ch_num, ch_text, model_id
                )

                if not blocks:
                    logger.warning(f"Chapter {ch_num}: No blocks extracted")
                    return ch_key, []

                raw_blocks = []
                for block in blocks:
                    block_id = block.get("block_id", "?")
                    q_start = block.get("question_start", 0)
                    q_end = block.get("question_end", 0)
                    a_start = block.get("answer_start", 0)
                    a_end = block.get("answer_end", 0)

                    # Extract raw text WITH line numbers preserved
                    question_text_raw = extract_lines_by_range_mapped(
                        lines_with_images, q_start, q_end, line_mapping, preserve_line_numbers=True
                    ) if q_start > 0 else ""

                    answer_text_raw = extract_lines_by_range_mapped(
                        lines_with_images, a_start, a_end, line_mapping, preserve_line_numbers=True
                    ) if a_start > 0 else ""

                    # Calculate page numbers from line ranges
                    q_pages = get_pages_for_line_range(q_start, q_end, pages_with_lines)
                    a_pages = get_pages_for_line_range(a_start, a_end, pages_with_lines)

                    raw_blocks.append({
                        "block_id": f"ch{ch_num}_{block_id}",
                        "block_label": block_id,
                        "chapter": ch_num,
                        "question_start": q_start,
                        "question_end": q_end,
                        "answer_start": a_start,
                        "answer_end": a_end,
                        "question_pages": q_pages,
                        "answer_pages": a_pages,
                        "question_text_raw": question_text_raw,
                        "answer_text_raw": answer_text_raw,
                        "formatted": False
                    })

                logger.info(f"Chapter {ch_num}: Extracted {len(raw_blocks)} blocks with line numbers")
                return ch_key, raw_blocks

            # Extract chapters in parallel
            status_text.text(f"Extracting {total_chapters} chapters with {max_workers} parallel workers...")
            completed = 0
            in_progress = set()
            block_results = {}

            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_ch = {
                    executor.submit(extract_block_worker, ch_data): ch_data["ch_num"]
                    for ch_data in chapter_data
                }

                for ch_data in chapter_data[:max_workers]:
                    in_progress.add(ch_data["ch_num"])
                chapter_status.text(f"In progress: Ch {', Ch '.join(map(str, sorted(in_progress)))}")

                for future in as_completed(future_to_ch):
                    ch_num = future_to_ch[future]
                    try:
                        ch_key, raw_blocks = future.result()
                        block_results[ch_key] = raw_blocks
                        st.session_state.raw_blocks[ch_key] = raw_blocks
                        save_raw_blocks()
                    except Exception as e:
                        logger.error(f"Chapter {ch_num}: Extraction failed - {e}")
                        block_results[f"ch{ch_num}"] = []

                    completed += 1
                    in_progress.discard(ch_num)

                    if completed + len(in_progress) <= total_chapters:
                        for ch_data in chapter_data:
                            if ch_data["ch_num"] not in in_progress and f"ch{ch_data['ch_num']}" not in block_results:
                                in_progress.add(ch_data["ch_num"])
                                break

                    progress_bar.progress(completed / total_chapters)
                    status_text.text(f"Completed {completed}/{total_chapters} chapters...")
                    if in_progress:
                        chapter_status.text(f"In progress: Ch {', Ch '.join(map(str, sorted(in_progress)))}")
                    else:
                        chapter_status.empty()

            st.session_state.raw_blocks = block_results
            save_raw_blocks()
            save_cost_tracking(get_output_dir())
            status_text.text("Done!")
            chapter_status.empty()

            total_blocks = sum(len(bs) for bs in block_results.values())
            st.success(f"Extracted {total_blocks} blocks from {total_chapters} chapters")
            render_cost_metrics("identify_blocks")
            play_completion_sound()
            st.info("**Next:** Go to **Step 4: Format Questions** to format the raw blocks.")
            st.rerun()

    # Display raw blocks if available
    raw_blocks = st.session_state.get("raw_blocks", {})
    if raw_blocks:
        st.markdown("---")
        st.subheader("Raw Extracted Blocks")

        total_blocks = sum(len(bs) for bs in raw_blocks.values())
        st.success(f"Total: {total_blocks} blocks across {len(raw_blocks)} chapters")
        st.info("**Next:** Go to **Step 4: Format Questions** to format these into structured data.")

        # Chapter selector for preview
        ch_options = sort_chapter_keys(raw_blocks.keys())
        if ch_options:
            selected_ch = st.selectbox("Preview chapter:", ch_options, key="raw_preview_ch")

            if selected_ch and selected_ch in raw_blocks:
                ch_raw = raw_blocks[selected_ch]
                st.caption(f"{len(ch_raw)} blocks in {selected_ch}")

                for block in ch_raw:
                    block_label = block.get("block_label", "?")
                    q_preview = block.get("question_text_raw", "")[:100] + "..." if len(block.get("question_text_raw", "")) > 100 else block.get("question_text_raw", "")
                    formatted_indicator = " [formatted]" if block.get("formatted", False) else ""

                    with st.expander(f"Block {block_label}{formatted_indicator}: {q_preview}"):
                        # Get page numbers - use stored values if available, otherwise calculate
                        q_pages = block.get('question_pages') or get_pages_for_line_range(block['question_start'], block['question_end'], st.session_state.pages)
                        a_pages = block.get('answer_pages') or get_pages_for_line_range(block['answer_start'], block['answer_end'], st.session_state.pages)
                        q_pages_str = f"{q_pages[0]}-{q_pages[-1]}" if len(q_pages) > 1 else str(q_pages[0]) if q_pages else "?"
                        a_pages_str = f"{a_pages[0]}-{a_pages[-1]}" if len(a_pages) > 1 else str(a_pages[0]) if a_pages else "?"

                        st.markdown(f"**Lines:** Q={block['question_start']}-{block['question_end']}, A={block['answer_start']}-{block['answer_end']} | **Pages:** Q={q_pages_str}, A={a_pages_str}")

                        st.markdown("**Question Text (raw with line numbers):**")
                        st.text_area("", block.get("question_text_raw", ""), height=200, key=f"raw_q_{block['block_id']}", disabled=True)

                        if block.get("answer_text_raw"):
                            st.markdown("**Answer Text (raw with line numbers):**")
                            st.text_area("", block.get("answer_text_raw", ""), height=200, key=f"raw_a_{block['block_id']}", disabled=True)


# =============================================================================
# Step 4: Format Questions
# =============================================================================

def render_format_step():
    """Render question formatting step - formats raw Q&A pairs using parallel LLM calls."""
    col1, col2 = st.columns([4, 1])
    with col1:
        st.header("Step 4: Format Questions")
    with col2:
        if st.button("Reset To This Step", key="clear_format", type="secondary", help="Clear this step and all subsequent steps"):
            clear_step_data("format")
            st.success("Format and subsequent data cleared")
            st.rerun()

    # Show cost from last completed step
    display_last_step_cost()

    st.markdown("""
    This step takes the raw Q&A blocks and formats them into structured data using parallel LLM calls.
    Each block is processed individually, extracting context, sub-questions, choices, correct answers, and explanations.
    """)

    # Load raw_blocks if not in session state
    raw_blocks_file = get_raw_blocks_file()
    if not st.session_state.raw_blocks and os.path.exists(raw_blocks_file):
        with open(raw_blocks_file, "r") as f:
            st.session_state.raw_blocks = json.load(f)

    raw_blocks = st.session_state.get("raw_blocks", {})

    if not raw_blocks:
        st.warning("Please extract raw questions first (Step 3)")
        return

    client = get_anthropic_client()
    if not client:
        st.error("ANTHROPIC_API_KEY not set. Please configure your .env file.")
        return

    # Initialize logger
    output_dir = get_output_dir()
    logger = get_extraction_logger(output_dir)

    total_blocks = sum(len(bs) for bs in raw_blocks.values())
    total_formatted = sum(len(qs) for qs in st.session_state.questions.values())

    st.success(f"**{total_blocks} raw blocks** with line numbers preserved")
    st.info(f"Formatted questions: {total_formatted}")

    # Chapter selector
    if not st.session_state.chapters:
        st.warning("No chapters available. Please extract chapters first (Step 2)")
        return

    chapter_options = [f"Ch{ch['chapter_number']}: {ch['title'][:40]}..."
                      for ch in st.session_state.chapters]
    selected_ch_idx = st.selectbox("Select chapter:",
                                    range(len(chapter_options)),
                                    format_func=lambda x: chapter_options[x],
                                    key="format_ch_selector")

    selected_ch = st.session_state.chapters[selected_ch_idx]
    selected_ch_num = selected_ch["chapter_number"]
    selected_ch_key = f"ch{selected_ch_num}"

    # Count blocks for selected chapter
    ch_raw_blocks = raw_blocks.get(selected_ch_key, [])
    ch_block_count = len(ch_raw_blocks)

    # Model selection and parallel workers
    model_col, workers_col, btn_col1, btn_col2 = st.columns([2, 1.5, 2, 2])

    with model_col:
        model_options = get_model_options()
        current_idx = model_options.index(st.session_state.selected_model) if st.session_state.selected_model in model_options else 0
        selected_model = st.selectbox("Model:", model_options, index=current_idx, key="format_model")
        if selected_model != st.session_state.selected_model:
            st.session_state.selected_model = selected_model
            save_settings()

    with workers_col:
        max_workers = st.number_input(
            "Parallel workers:",
            min_value=1,
            max_value=100,
            value=50,
            help="Tier 1: use 5-10 | Tier 2+: use 20-50 | Tier 4: use 50-100",
            key="format_workers"
        )

    # Capture model_id before threads start (session state not accessible in worker threads)
    model_id = get_selected_model_id()

    # Helper function to format a single block (v2 format)
    def format_single_block(item):
        """Format a raw block using format_raw_block_llm and convert to questions."""
        import re
        ch_key, block = item
        ch_num = block["chapter"]
        block_id = block.get("block_label", block.get("block_id", "?"))

        # Strip line number markers before formatting (they were for traceability during extraction)
        question_text = block.get("question_text_raw", "")
        answer_text = block.get("answer_text_raw", "")

        # Remove [LINE:NNNN] markers
        question_text = re.sub(r'\[LINE:\d+\]\s*', '', question_text)
        answer_text = re.sub(r'\[LINE:\d+\]\s*', '', answer_text)

        # Call the format_raw_block_llm function with cleaned text
        formatted_block = format_raw_block_llm(
            client,
            block_id,
            question_text,
            answer_text,
            model_id,
            ch_num
        )

        # Convert formatted block to questions format
        questions = []

        context = formatted_block.get("context", {})
        context_text = context.get("text", "")
        context_images = context.get("image_files", [])

        # Get shared discussion
        shared = formatted_block.get("shared_discussion", {})
        shared_text = shared.get("full_text", "")
        shared_answer_images = shared.get("image_files", [])  # Images from answer section
        if not shared_text:
            # Build from components
            parts = []
            if shared.get("imaging_findings"):
                parts.append(f"**Imaging Findings:** {shared['imaging_findings']}")
            if shared.get("discussion"):
                parts.append(f"**Discussion:** {shared['discussion']}")
            if shared.get("differential_diagnosis"):
                parts.append(f"**Differential Diagnosis:** {shared['differential_diagnosis']}")
            if shared.get("references"):
                parts.append("**References:**\n" + "\n".join(f"- {r}" for r in shared["references"]))
            shared_text = "\n\n".join(parts)

        # Process sub-questions
        sub_questions = formatted_block.get("sub_questions", [])

        # If there's context but no sub-questions, it's context-only
        if context_text and not sub_questions:
            questions.append({
                "full_id": f"ch{ch_num}_{block_id}",
                "local_id": block_id,
                "chapter": ch_num,
                "question_start": block.get("question_start", 0),
                "question_end": block.get("question_end", 0),
                "answer_start": block.get("answer_start", 0),
                "answer_end": block.get("answer_end", 0),
                "correct_letter": "",
                "image_files": context_images,
                "question_text": context_text,
                "answer_text": "",
                "is_context_only": True,
                "block_id": block.get("block_id", f"ch{ch_num}_block_{block_id}")
            })

        # Add sub-questions
        for sq in sub_questions:
            local_id = sq.get("local_id", "?")
            full_id = f"ch{ch_num}_{local_id}"

            # Combine context with question text for sub-questions
            q_text = sq.get("question_text", "")
            if context_text and local_id != block_id:
                q_text = context_text + "\n\n" + q_text

            # Build explanation - use exactly what the LLM returns
            explanation = sq.get("explanation", "")

            # Combine question-specific images with context images
            # Context images apply to ALL sub-questions in the block
            all_images = list(sq.get("image_files", []))
            for img in context_images:
                if img not in all_images:
                    all_images.append(img)

            # Combine sub-question answer images with shared answer images
            all_answer_images = list(sq.get("answer_image_files", []))
            for img in shared_answer_images:
                if img not in all_answer_images:
                    all_answer_images.append(img)

            # Get page numbers from LLM response
            question_pages = sq.get("question_pages", [])
            answer_pages = sq.get("answer_pages", [])

            questions.append({
                "id": full_id,
                "full_id": full_id,
                "local_id": local_id,
                "chapter": ch_num,
                "question_start": block.get("question_start", 0),
                "question_end": block.get("question_end", 0),
                "answer_start": block.get("answer_start", 0),
                "answer_end": block.get("answer_end", 0),
                "correct_letter": sq.get("correct_answer", ""),
                "image_files": all_images,
                "answer_image_files": all_answer_images,
                "question_text": q_text,
                "answer_text": explanation,
                "is_context_only": False,
                "block_id": block.get("block_id", f"ch{ch_num}_block_{block_id}"),
                "choices": sq.get("choices", {}),
                "text": q_text,  # For compatibility
                "correct_answer": sq.get("correct_answer", ""),
                "explanation": explanation,
                "question_pages": question_pages,
                "question_page": question_pages[0] if question_pages else None,
                "answer_pages": answer_pages,
                "answer_page": answer_pages[0] if answer_pages else None
            })

        return ch_key, questions, formatted_block

    with btn_col1:
        format_chapter = st.button(
            f"Format Chapter {selected_ch_num} ({ch_block_count} blocks)",
            type="primary",
            key="format_chapter_btn",
            disabled=ch_block_count == 0
        )

    with btn_col2:
        format_all = st.button("Format ALL", type="secondary", key="format_all_btn")

    # Single chapter formatting logic
    if format_chapter:
        if ch_block_count == 0:
            st.warning(f"No raw blocks in Chapter {selected_ch_num}")
        else:
            progress_bar = st.progress(0)
            status_text = st.empty()

            # Block-based formatting
            ch_blocks = [(selected_ch_key, block) for block in ch_raw_blocks]
            total = len(ch_blocks)
            status_text.text(f"Formatting {total} blocks from Chapter {selected_ch_num}...")

            formatted_list = []
            completed = 0

            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {executor.submit(format_single_block, item): item for item in ch_blocks}

                for future in as_completed(futures):
                    try:
                        ch_key, questions, _ = future.result()
                        formatted_list.extend(questions)
                    except Exception as e:
                        ch_key, block = futures[future]
                        logger.error(f"Error formatting block {block.get('block_id', '?')}: {e}")
                        # Add placeholder
                        formatted_list.append({
                            "full_id": block.get("block_id", "?"),
                            "local_id": block.get("block_label", "?"),
                            "text": block.get("question_text_raw", ""),
                            "choices": {},
                            "correct_answer": "",
                            "explanation": block.get("answer_text_raw", ""),
                            "image_files": [],
                            "error": str(e)
                        })

                    completed += 1
                    progress_bar.progress(completed / total)
                    status_text.text(f"Chapter {selected_ch_num}: {completed}/{total} blocks formatted...")

            # Sort and save
            formatted_list.sort(key=lambda q: question_sort_key(q["full_id"]))
            st.session_state.questions[selected_ch_key] = formatted_list
            save_questions()

            # Build image assignments for this chapter with block-aware context inheritance
            chapter_dict = {selected_ch_key: formatted_list}
            new_assignments = build_block_aware_image_assignments(chapter_dict)
            st.session_state.image_assignments.update(new_assignments)
            st.session_state.questions[selected_ch_key] = formatted_list  # context_from was set by helper
            save_questions()
            save_image_assignments()

            save_cost_tracking(get_output_dir())
            status_text.text("Done!")
            st.success(f"Formatted {total} blocks from Chapter {selected_ch_num}")
            render_cost_metrics("format_blocks")
            play_completion_sound()
            st.rerun()

    # All chapters formatting logic
    if format_all:
        progress_bar = st.progress(0)
        status_text = st.empty()

        # Block-based formatting for all chapters
        all_blocks = []
        for ch_key, ch_blocks in raw_blocks.items():
            for block in ch_blocks:
                all_blocks.append((ch_key, block))

        total = len(all_blocks)
        status_text.text(f"Formatting {total} blocks with {max_workers} parallel workers...")

        formatted_by_chapter = {}
        completed = 0

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(format_single_block, item): item for item in all_blocks}

            for future in as_completed(futures):
                try:
                    ch_key, questions, _ = future.result()
                    if ch_key not in formatted_by_chapter:
                        formatted_by_chapter[ch_key] = []
                    formatted_by_chapter[ch_key].extend(questions)
                except Exception as e:
                    ch_key, block = futures[future]
                    logger.error(f"Error formatting block {block.get('block_id', '?')}: {e}")
                    if ch_key not in formatted_by_chapter:
                        formatted_by_chapter[ch_key] = []
                    formatted_by_chapter[ch_key].append({
                        "full_id": block.get("block_id", "?"),
                        "local_id": block.get("block_label", "?"),
                        "text": block.get("question_text_raw", ""),
                        "choices": {},
                        "correct_answer": "",
                        "explanation": block.get("answer_text_raw", ""),
                        "image_files": [],
                        "error": str(e)
                    })

                completed += 1
                progress_bar.progress(completed / total)
                status_text.text(f"Formatted {completed}/{total} blocks...")

                # Save incrementally every 5 blocks
                if completed % 5 == 0:
                    st.session_state.questions = formatted_by_chapter
                    save_questions()

        # Sort questions within each chapter
        for ch_key in formatted_by_chapter:
            formatted_by_chapter[ch_key].sort(key=lambda q: question_sort_key(q["full_id"]))

        st.session_state.questions = formatted_by_chapter
        save_questions()

        # Build image assignments with block-aware context inheritance
        st.session_state.image_assignments = build_block_aware_image_assignments(formatted_by_chapter)
        st.session_state.questions = formatted_by_chapter  # context_from was set by helper
        save_questions()
        save_image_assignments()

        # If raw_blocks exist, save them as question_blocks (ready for block-based generation)
        if raw_blocks:
            st.session_state.question_blocks = raw_blocks
            save_question_blocks()
            logger.info(f"Saved {sum(len(bs) for bs in raw_blocks.values())} blocks to question_blocks.json")

        save_cost_tracking(get_output_dir())
        status_text.text("Done!")
        st.success(f"Formatted {total} blocks")
        render_cost_metrics("format_blocks")
        play_completion_sound()
        st.info("**Next:** Go to **Step 5: QC Questions** to review and approve extracted questions.")
        st.rerun()

    # Display formatted questions if available
    if st.session_state.questions:
        st.markdown("---")
        st.subheader("Formatted Questions")

        ch_options = sort_chapter_keys(st.session_state.questions.keys())
        if ch_options:
            selected_ch = st.selectbox("Preview chapter:", ch_options, key="format_preview_ch")

            if selected_ch and selected_ch in st.session_state.questions:
                questions = st.session_state.questions[selected_ch]
                st.caption(f"{len(questions)} formatted questions in {selected_ch}")

                for q in questions:
                    # Get question images and answer images from the question's arrays
                    q_image_files = set(q.get("image_files", []))
                    a_image_files = set(q.get("answer_image_files", []))
                    q_images = [img for img in st.session_state.images if img["filename"] in q_image_files]
                    a_images = [img for img in st.session_state.images if img["filename"] in a_image_files]

                    total_imgs = len(q_images) + len(a_images)
                    img_indicator = f" [{total_imgs} img]" if total_imgs > 0 else ""
                    error_indicator = " [ERROR]" if q.get("error") else ""
                    display_text = q['text'][:70] + "..." if len(q['text']) > 70 else q['text']

                    with st.expander(f"Q{q['local_id']}{img_indicator}{error_indicator}: {display_text}"):
                        if q.get("error"):
                            st.error(f"Formatting error: {q['error']}")

                        col1, col2 = st.columns([2, 1])

                        with col1:
                            st.markdown(f"**Question:** {q['text']}")

                            if q.get("choices"):
                                st.markdown("**Choices:**")
                                for letter, choice in q.get("choices", {}).items():
                                    st.markdown(f"- {letter}: {choice}")
                                st.markdown(f"**Correct Answer:** {q.get('correct_answer', 'N/A')}")

                            if q.get("explanation"):
                                st.markdown(f"**Explanation:** {q.get('explanation', '')}")

                        with col2:
                            if q_images:
                                st.markdown("**Question Images:**")
                                for img in q_images:
                                    if os.path.exists(img["filepath"]):
                                        st.image(img["filepath"], caption=f"Page {img['page']}", width=200)

                            if a_images:
                                st.markdown("**Answer Images:**")
                                for img in a_images:
                                    if os.path.exists(img["filepath"]):
                                        st.image(img["filepath"], caption=f"Page {img['page']} (answer)", width=200)


# =============================================================================
# Step 5: QC Questions
# =============================================================================

def render_qc_step():
    """Render QC review step."""
    col1, col2 = st.columns([4, 1])
    with col1:
        st.header("Step 5: QC Questions")
    with col2:
        if st.button("Reset To This Step", key="clear_qc", type="secondary", help="Clear this step and all subsequent steps"):
            clear_step_data("qc")
            st.success("QC and export data cleared")
            st.rerun()

    # Use formatted questions from Step 4
    questions_source = st.session_state.questions

    if not questions_source:
        st.warning("Please format questions first (Step 4)")
        return

    all_questions = []
    for ch_key, questions in questions_source.items():
        for q in questions:
            all_questions.append((ch_key, q))

    all_questions.sort(key=lambda x: question_sort_key(x[1]["full_id"]))

    if not all_questions:
        st.warning("No questions to review")
        return

    reviewed = st.session_state.qc_progress.get("reviewed", {})
    total = len(all_questions)
    reviewed_count = len(reviewed)

    st.progress(reviewed_count / total if total > 0 else 0)
    st.caption(f"Progress: {reviewed_count}/{total} questions reviewed")

    col1, col2, col3 = st.columns(3)
    with col1:
        filter_option = st.radio("Show:", ["All", "Unreviewed only", "Reviewed only"], horizontal=True)
    with col2:
        chapter_filter = st.selectbox("Filter by chapter:",
                                       ["All chapters"] + sort_chapter_keys(questions_source.keys()))
    with col3:
        hide_context = st.checkbox("Hide context-only entries", value=True)

    filtered_questions = []
    for ch_key, q in all_questions:
        q_id = q["full_id"]
        is_reviewed = q_id in reviewed

        if filter_option == "Unreviewed only" and is_reviewed:
            continue
        if filter_option == "Reviewed only" and not is_reviewed:
            continue
        if chapter_filter != "All chapters" and ch_key != chapter_filter:
            continue
        if hide_context and q.get("is_context_only"):
            continue

        filtered_questions.append((ch_key, q))

    st.caption(f"Showing {len(filtered_questions)} questions")

    if filtered_questions:
        def format_question_option(q):
            prefix = ""
            if q.get("is_context_only"):
                prefix = "[CTX] "
            elif q.get("context"):
                prefix = "[+ctx] "
            return f"{prefix}{q['full_id']}: {q['text'][:50]}..."

        question_options = [format_question_option(q) for _, q in filtered_questions]

        if st.session_state.qc_selected_idx >= len(filtered_questions):
            st.session_state.qc_selected_idx = 0

        # Define callbacks for navigation (run before widgets render)
        def go_previous():
            if st.session_state.qc_selected_idx > 0:
                st.session_state.qc_selected_idx -= 1
                # Delete the selectbox's stored state so it uses the new index
                if "qc_question_selector" in st.session_state:
                    del st.session_state.qc_question_selector
                save_settings()

        def go_next():
            if st.session_state.qc_selected_idx < len(filtered_questions) - 1:
                st.session_state.qc_selected_idx += 1
                # Delete the selectbox's stored state so it uses the new index
                if "qc_question_selector" in st.session_state:
                    del st.session_state.qc_question_selector
                save_settings()

        selected_idx = st.selectbox(
            "Select question:",
            range(len(question_options)),
            index=st.session_state.qc_selected_idx,
            format_func=lambda x: question_options[x],
            key="qc_question_selector"
        )
        if selected_idx != st.session_state.qc_selected_idx:
            st.session_state.qc_selected_idx = selected_idx
            save_settings()

        selected_idx = st.session_state.qc_selected_idx
        if selected_idx is not None and selected_idx < len(filtered_questions):
            ch_key, q = filtered_questions[selected_idx]
            q_id = q["full_id"]

            # Get review status for this question
            is_approved = q_id in reviewed and reviewed[q_id].get("status") == "approved"
            is_flagged = q_id in reviewed and reviewed[q_id].get("status") == "flagged"

            # Define callbacks for QC actions
            def approve_and_next():
                reviewed[q_id] = {"status": "approved", "timestamp": datetime.now().isoformat()}
                st.session_state.qc_progress["reviewed"] = reviewed
                save_qc_progress()
                if st.session_state.qc_selected_idx < len(filtered_questions) - 1:
                    st.session_state.qc_selected_idx += 1
                    if "qc_question_selector" in st.session_state:
                        del st.session_state.qc_question_selector
                    save_settings()

            def flag_issue():
                reviewed[q_id] = {"status": "flagged", "timestamp": datetime.now().isoformat()}
                st.session_state.qc_progress["reviewed"] = reviewed
                save_qc_progress()

            def unapprove():
                reviewed.pop(q_id, None)
                st.session_state.qc_progress["reviewed"] = reviewed
                save_qc_progress()

            # Navigation and action buttons in a single row (fixed position)
            # Column widths tuned to prevent text wrapping
            col1, col2, col3, col4, col5, col6 = st.columns([1, 1, 1.5, 1.2, 1.3, 1.2])

            with col1:
                st.button("Previous", disabled=(st.session_state.qc_selected_idx <= 0), on_click=go_previous)

            with col2:
                st.button("Next", disabled=(st.session_state.qc_selected_idx >= len(filtered_questions) - 1), on_click=go_next)

            with col3:
                st.button("Approve & Next", type="primary", disabled=is_approved, on_click=approve_and_next)

            with col4:
                st.button("Flag Issue", disabled=is_flagged, on_click=flag_issue)

            with col5:
                if st.button("Detect Pages", help="Detect PDF pages for each question"):
                    with st.spinner("Detecting page numbers..."):
                        add_page_numbers_to_questions(
                            st.session_state.questions,
                            st.session_state.pages,
                            st.session_state.chapters
                        )
                        save_questions()
                        st.rerun()

            with col6:
                if is_approved or is_flagged:
                    st.button("Unapprove", on_click=unapprove)

            st.markdown("---")

            left_col, right_col = st.columns([1, 1])

            with left_col:
                # Status indicator with color
                if is_approved:
                    status_html = '<span style="color: green;">(approved)</span>'
                elif is_flagged:
                    status_html = '<span style="color: orange;">(flagged)</span>'
                else:
                    status_html = '<span style="color: red;">(pending)</span>'
                st.markdown(f"### Question {q['local_id']} {status_html}", unsafe_allow_html=True)

                if q.get("is_context_only"):
                    st.warning("**CONTEXT ONLY** - This entry provides context for sub-questions and will not be exported to Anki.")
                    st.markdown(f"**Context Text:**\n\n{q['text']}")
                else:
                    if q.get("context"):
                        st.info(f"**Context (from Q{q.get('context_question_id', '?')}):**\n\n{q['context']}")

                    st.markdown(f"**Question:** {q['text']}")

                    if q.get("choices"):
                        st.markdown("**Choices:**")
                        for letter, choice in q.get("choices", {}).items():
                            if letter == q.get("correct_answer"):
                                st.markdown(f"- **{letter}: {choice}** (correct)")
                            else:
                                st.markdown(f"- {letter}: {choice}")
                    else:
                        st.caption("No answer choices")

                    st.markdown(f"**Explanation:** {q.get('explanation', 'N/A')}")

                # Block context panel - show if question belongs to a block
                question_blocks = st.session_state.get("question_blocks", {})
                block_id = q.get("block_id")
                if block_id and question_blocks:
                    # Find the source block
                    source_block = None
                    for ch_blk_key, ch_blocks in question_blocks.items():
                        for block in ch_blocks:
                            if block.get("block_id") == block_id or block.get("block_label") == block_id:
                                source_block = block
                                break
                        if source_block:
                            break

                    if source_block:
                        with st.expander("View Full Block Context", expanded=False):
                            st.markdown(f"**Block ID:** `{block_id}`")

                            # Show block context (clinical scenario)
                            if source_block.get("context_text"):
                                st.markdown("**Context (Clinical Scenario):**")
                                st.text_area("", source_block["context_text"], height=100, disabled=True, key=f"qc_block_ctx_{q_id}")

                            # Show all sub-questions in this block
                            sub_questions = source_block.get("sub_questions", [])
                            if len(sub_questions) > 1:
                                st.markdown(f"**Other questions in this block ({len(sub_questions)} total):**")
                                for sq in sub_questions:
                                    sq_local = sq.get("local_id", "")
                                    sq_text = sq.get("question_text", "")[:80]
                                    if sq_local != q.get("local_id"):
                                        st.markdown(f"- Q{sq_local}: {sq_text}...")

                            # Show shared answer/discussion
                            shared_answer = source_block.get("shared_answer_text", "")
                            if shared_answer:
                                st.markdown("**Shared Discussion (applies to all sub-questions):**")
                                st.text_area("", shared_answer, height=200, disabled=True, key=f"qc_block_shared_{q_id}")

            with right_col:
                question_images = get_images_for_question(q_id)
                answer_images = get_answer_images_for_question(q_id)

                ch_num = int(ch_key[2:])
                ch_start = next((c["start_page"] for c in st.session_state.chapters if c["chapter_number"] == ch_num), 1)
                ch_end = 9999
                for i, c in enumerate(st.session_state.chapters):
                    if c["chapter_number"] == ch_num and i + 1 < len(st.session_state.chapters):
                        ch_end = st.session_state.chapters[i + 1]["start_page"]

                # Define callback functions for image operations
                def remove_image(img_filename):
                    st.session_state.image_assignments.pop(img_filename, None)
                    save_image_assignments()
                    if "qc_question_selector" in st.session_state:
                        del st.session_state.qc_question_selector

                def assign_image(img_filename, question_id):
                    st.session_state.image_assignments[img_filename] = question_id
                    save_image_assignments()
                    if "qc_question_selector" in st.session_state:
                        del st.session_state.qc_question_selector

                # Create tabs for Images and PDF Pages
                images_tab, pdf_tab = st.tabs(["Images", "PDF Pages"])

                with images_tab:
                    # Question images (shown with the question)
                    if question_images:
                        st.subheader("Question Image(s)")
                        for img in question_images:
                            filepath = img["filepath"]
                            if os.path.exists(filepath):
                                st.image(filepath, caption=f"Page {img['page']} - {img['filename']}")

                                st.button("Remove Image", key=f"img_remove_{img['filename']}",
                                         on_click=remove_image, args=(img["filename"],))
                            else:
                                st.warning(f"Image not found: {filepath}")

                    # Answer images (shown in explanation only)
                    if answer_images:
                        st.subheader("Answer Image(s)")
                        st.caption("These images appear in the explanation only")
                        for img in answer_images:
                            filepath = img["filepath"]
                            if os.path.exists(filepath):
                                st.image(filepath, caption=f"Page {img['page']} - {img['filename']} (answer)")
                            else:
                                st.warning(f"Image not found: {filepath}")

                    # Expander for adding more images (if there are already some)
                    if question_images or answer_images:
                        with st.expander("Assign different/additional image"):
                            unassigned = [img for img in st.session_state.images
                                          if img["filename"] not in st.session_state.image_assignments
                                          and ch_start <= img["page"] < ch_end]
                            if unassigned:
                                for img in unassigned[:5]:
                                    filepath = img["filepath"]
                                    if os.path.exists(filepath):
                                        st.image(filepath, caption=f"Page {img['page']}", width=150)
                                        st.button(f"Add this image", key=f"add_{img['filename']}",
                                                 on_click=assign_image, args=(img["filename"], q_id))
                            else:
                                st.caption("No more unassigned images in this chapter")

                    else:
                        if q.get("has_image"):
                            st.subheader("Image Required")
                            st.warning("This question needs an image but none assigned yet.")
                        else:
                            st.subheader("No Image Assigned")
                            st.caption("No image currently linked to this question")

                        with st.expander("Manually assign an image" if not q.get("has_image") else "Select from chapter images", expanded=q.get("has_image", False)):
                            unassigned = [img for img in st.session_state.images
                                          if img["filename"] not in st.session_state.image_assignments
                                          and ch_start <= img["page"] < ch_end]

                            if unassigned:
                                for img in unassigned[:6]:
                                    filepath = img["filepath"]
                                    if os.path.exists(filepath):
                                        st.image(filepath, caption=f"Page {img['page']}", width=180)
                                        st.button(f"Assign", key=f"assign_{img['filename']}",
                                                 on_click=assign_image, args=(img["filename"], q_id))
                            else:
                                st.caption("No unassigned images in this chapter")

                with pdf_tab:
                    # Support both single page (legacy) and multi-page formats
                    question_pages = q.get("question_pages", [])
                    answer_pages = q.get("answer_pages", [])
                    # Fall back to single page if lists not available
                    if not question_pages and q.get("question_page"):
                        question_pages = [q.get("question_page")]
                    if not answer_pages and q.get("answer_page"):
                        answer_pages = [q.get("answer_page")]

                    pdf_path = st.session_state.get("pdf_path")

                    if not pdf_path or not os.path.exists(pdf_path):
                        st.warning("PDF file not available. Re-load the textbook in Step 1 to enable PDF preview.")
                    elif not question_pages and not answer_pages:
                        st.info("Page numbers not detected for this question.")
                        if st.button("Detect Page Numbers", key=f"detect_pages_{q_id}"):
                            with st.spinner("Detecting page numbers..."):
                                add_page_numbers_to_questions(
                                    st.session_state.questions,
                                    st.session_state.pages,
                                    st.session_state.chapters
                                )
                                save_questions()
                                st.rerun()
                    else:
                        # Show question pages and answer pages side by side
                        q_col, a_col = st.columns(2)

                        with q_col:
                            if len(question_pages) > 1:
                                st.markdown(f"**Question Pages ({len(question_pages)})**")
                            else:
                                st.markdown("**Question Page**")

                            if question_pages:
                                for page_num in question_pages:
                                    png_bytes = render_pdf_page(pdf_path, page_num, zoom=1.2)
                                    if png_bytes:
                                        st.image(png_bytes, caption=f"Page {page_num}")
                                    else:
                                        st.error(f"Failed to render page {page_num}")
                            else:
                                st.caption("Page not detected")

                        with a_col:
                            if len(answer_pages) > 1:
                                st.markdown(f"**Answer Pages ({len(answer_pages)})**")
                            else:
                                st.markdown("**Answer Page**")

                            if answer_pages:
                                for page_num in answer_pages:
                                    png_bytes = render_pdf_page(pdf_path, page_num, zoom=1.2)
                                    if png_bytes:
                                        st.image(png_bytes, caption=f"Page {page_num}")
                                    else:
                                        st.error(f"Failed to render page {page_num}")
                            else:
                                st.caption("Page not detected")


# =============================================================================
# Anki Deck Generation Helper
# =============================================================================

def generate_anki_deck(book_name: str, questions: dict, chapters: list, image_assignments: dict,
                       images: list, include_images: bool, only_approved: bool, qc_progress: dict,
                       generated_questions: dict = None, include_generated: bool = False,
                       export_selections: dict = None) -> str:
    """Generate Anki deck and return path to .apkg file.

    Args:
        book_name: Name for the deck
        questions: Dict of chapter questions
        chapters: List of chapter info
        image_assignments: Dict mapping image files to question IDs
        images: List of image metadata
        include_images: Whether to include images in cards
        only_approved: Only export QC-approved questions
        qc_progress: QC progress dict
        generated_questions: Optional dict of generated cloze cards
        include_generated: Whether to include generated cloze cards
        export_selections: Dict of chapter/type selections (e.g., {"ch1_extracted": True, "ch1_generated": False})
    """
    # Default to all selected if no selections provided
    if export_selections is None:
        export_selections = {}
    import genanki
    import hashlib

    # Generate stable IDs based on name
    def stable_id(name: str) -> int:
        return int(hashlib.md5(name.encode()).hexdigest()[:8], 16)

    model_id = stable_id(f"{book_name}_model")
    cloze_model_id = stable_id(f"{book_name}_cloze_model")

    # Create card model (template)
    model = genanki.Model(
        model_id,
        f'{book_name} Model',
        fields=[
            {'name': 'Question'},
            {'name': 'Choices'},
            {'name': 'Answer'},
            {'name': 'Explanation'},
            {'name': 'Image'},
            {'name': 'AnswerImage'},
            {'name': 'Chapter'},
            {'name': 'Source'},
        ],
        templates=[
            {
                'name': 'Card 1',
                'qfmt': '''
                    <div class="card-content">
                        <div class="question">{{Question}}</div>
                        {{#Image}}<div class="image">{{Image}}</div>{{/Image}}
                        <div class="choices">{{Choices}}</div>
                    </div>
                ''',
                'afmt': '''
                    <div class="card-content">
                        <div class="question">{{Question}}</div>
                        {{#Image}}<div class="image">{{Image}}</div>{{/Image}}
                        <div class="choices">{{Choices}}</div>
                        <hr class="answer-divider">
                        <div class="answer">{{Answer}}</div>
                        <div class="explanation">{{Explanation}}</div>
                        {{#AnswerImage}}<div class="image answer-image">{{AnswerImage}}</div>{{/AnswerImage}}
                        {{#Source}}<div class="source-ref">{{Source}}</div>{{/Source}}
                    </div>
                ''',
            },
        ],
        css='''
            .card {
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
                font-size: 17px;
                line-height: 1.5;
                color: #2c3e50;
                background: #fafafa;
            }
            .card-content {
                max-width: 650px;
                margin: 0 auto;
                padding: 25px;
                text-align: center;
            }
            .question {
                font-size: 18px;
                font-weight: 500;
                margin-bottom: 20px;
                text-align: left;
                line-height: 1.6;
            }
            .image {
                margin: 20px 0;
            }
            .image img {
                max-width: 100%;
                height: auto;
                border-radius: 8px;
                box-shadow: 0 2px 12px rgba(0,0,0,0.12);
            }
            .choices {
                text-align: left;
                margin: 20px 0;
            }
            .choice {
                margin: 10px 0;
                padding: 10px 14px;
                background: #ffffff;
                color: #2c3e50;
                border-radius: 6px;
                border: 1px solid #e1e5e9;
                border-left: 4px solid #3498db;
            }
            .choice-letter {
                font-weight: 700;
                color: #3498db;
                margin-right: 10px;
            }
            .answer-divider {
                border: none;
                border-top: 2px solid #e1e5e9;
                margin: 25px 0;
            }
            .answer {
                font-size: 20px;
                font-weight: 600;
                color: #27ae60;
                margin: 20px 0;
                padding: 12px 20px;
                background: linear-gradient(135deg, #d5f4e0 0%, #c8f0d8 100%);
                border-radius: 8px;
                display: inline-block;
            }
            .explanation {
                text-align: left;
                margin-top: 20px;
                padding: 18px;
                background: #ffffff;
                border-radius: 8px;
                border: 1px solid #e1e5e9;
                line-height: 1.7;
                color: #34495e !important;
            }
            .source-ref {
                margin-top: 20px;
                padding: 10px 14px;
                background: #f8f9fa;
                border-radius: 6px;
                font-size: 13px;
                color: #6c757d !important;
                text-align: left;
            }
            .source-ref-label {
                font-weight: 600;
                color: #495057 !important;
            }
        '''
    )

    # Create cloze model for generated cards
    cloze_model = genanki.Model(
        cloze_model_id,
        f'{book_name} Cloze Model',
        model_type=genanki.Model.CLOZE,
        fields=[
            {'name': 'Text'},
            {'name': 'Extra'},
            {'name': 'Source'},
        ],
        templates=[
            {
                'name': 'Cloze Card',
                'qfmt': '''
                    <div class="cloze-card">
                        <div class="cloze-text">{{cloze:Text}}</div>
                    </div>
                ''',
                'afmt': '''
                    <div class="cloze-card">
                        <div class="cloze-text">{{cloze:Text}}</div>
                        {{#Extra}}
                        <hr class="cloze-divider">
                        <div class="cloze-extra">{{Extra}}</div>
                        {{/Extra}}
                        {{#Source}}<div class="cloze-source">{{Source}}</div>{{/Source}}
                    </div>
                ''',
            },
        ],
        css='''
            .card {
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
                font-size: 18px;
                line-height: 1.6;
                color: #2c3e50;
                background: #fafafa;
            }
            .cloze-card {
                max-width: 650px;
                margin: 0 auto;
                padding: 25px;
            }
            .cloze-text {
                font-size: 18px;
                line-height: 1.7;
                text-align: left;
            }
            .cloze {
                font-weight: 700;
                color: #3498db;
            }
            .cloze-divider {
                border: none;
                border-top: 2px solid #e1e5e9;
                margin: 20px 0;
            }
            .cloze-extra {
                font-size: 15px;
                color: #5d6d7e;
                text-align: left;
                padding: 12px;
                background: #f8f9fa;
                border-radius: 6px;
            }
            .cloze-source {
                margin-top: 20px;
                padding: 14px;
                background: #f8f9fa;
                border: 1px solid #e1e5e9;
                border-radius: 6px;
                font-size: 13px;
                color: #6c757d;
                text-align: left;
                white-space: pre-wrap;
                line-height: 1.5;
            }
            .cloze-source-label {
                font-weight: 600;
                color: #495057;
                display: block;
                margin-bottom: 10px;
                border-bottom: 1px solid #e1e5e9;
                padding-bottom: 8px;
            }
            .cloze-source-text {
                font-size: 12px;
                color: #6c757d;
                max-height: 300px;
                overflow-y: auto;
            }
        '''
    )

    # Build image lookup
    image_lookup = {img['filename']: img for img in images}

    # Track media files to include
    media_files = []

    # Collect all chapter decks
    all_decks = []

    # Create sub-decks for each chapter
    reviewed = qc_progress.get("reviewed", {})

    for ch in chapters:
        ch_num = ch['chapter_number']
        ch_key = f"ch{ch_num}"
        ch_title = ch.get('title', f'Chapter {ch_num}')
        ch_questions = questions.get(ch_key, [])

        if not ch_questions:
            continue

        # Skip if this chapter's extracted cards are not selected
        # Default to True if no selection exists (backwards compatibility)
        if not export_selections.get(f"{ch_key}_extracted", True):
            continue

        # Create chapter sub-deck for extracted questions
        # Zero-pad chapter number for proper sorting (01, 02, ... 10, 11)
        ch_deck_name = f"{book_name}::{int(ch_num):02d}. {ch_title}::Extracted"
        ch_deck_id = stable_id(ch_deck_name)
        ch_deck = genanki.Deck(ch_deck_id, ch_deck_name)

        for q in ch_questions:
            q_id = q['full_id']

            # Skip context-only questions
            if q.get('is_context_only'):
                continue

            # Skip if only approved and not approved
            if only_approved:
                review_status = reviewed.get(q_id, {}).get('status')
                if review_status != 'approved':
                    continue

            # Build question text (include context if merged)
            q_text = q.get('text', '')

            # Build choices HTML with styled formatting
            choices = q.get('choices', {})
            if choices:
                choice_items = []
                for letter, text in sorted(choices.items()):
                    choice_items.append(
                        f'<div class="choice"><span class="choice-letter">{letter}.</span>{text}</div>'
                    )
                choices_html = ''.join(choice_items)
            else:
                choices_html = ''

            # Get correct answer
            correct = q.get('correct_answer', '')

            # Get explanation - include shared discussion from block if available
            explanation = q.get('explanation', '')

            # If question has a block_id, try to append shared discussion
            block_id = q.get('block_id')
            if block_id:
                question_blocks = st.session_state.get("question_blocks", {})
                if question_blocks:
                    for ch_key_b, ch_blocks in question_blocks.items():
                        for block in ch_blocks:
                            if block.get("block_id") == block_id or block.get("block_label") == block_id:
                                shared_answer = block.get('shared_answer_text', '')
                                if shared_answer and shared_answer not in explanation:
                                    explanation = explanation + "\n\n<b>Shared Discussion:</b>\n" + shared_answer
                                break

            # Handle images - question images and answer images separately
            image_html = ''
            answer_image_html = ''
            if include_images:
                # Question images (shown with the question)
                question_imgs = set(q.get('image_files', []))
                for img_fname in question_imgs:
                    if img_fname in image_lookup:
                        img_data = image_lookup[img_fname]
                        filepath = img_data.get('filepath', '')
                        if os.path.exists(filepath):
                            media_files.append(filepath)
                            image_html += f'<img src="{img_fname}">'

                # Answer images (shown only in explanation)
                answer_imgs = set(q.get('answer_image_files', []))
                for img_fname in answer_imgs:
                    if img_fname in image_lookup:
                        img_data = image_lookup[img_fname]
                        filepath = img_data.get('filepath', '')
                        if os.path.exists(filepath):
                            if filepath not in media_files:  # Avoid duplicates
                                media_files.append(filepath)
                            answer_image_html += f'<img src="{img_fname}">'

            # Build source reference
            local_id = q.get('local_id', q_id.split('_')[-1])
            source_parts = [f'<span class="source-ref-label">Source:</span> Chapter {ch_num} ({ch_title}), Question {local_id}']

            # Add page numbers if available
            question_page = q.get('question_page')
            answer_page = q.get('answer_page')
            if question_page and answer_page:
                source_parts.append(f'(Q: p.{question_page}, A: p.{answer_page})')
            elif question_page:
                source_parts.append(f'(p. {question_page})')
            elif answer_page:
                source_parts.append(f'(A: p.{answer_page})')

            if q.get('context_from'):
                context_local_id = q['context_from'].split('_')[-1]
                source_parts.append(f'[context from Q{context_local_id}]')

            source_ref = ' '.join(source_parts)

            # Create note
            note = genanki.Note(
                model=model,
                fields=[q_text, choices_html, correct, explanation, image_html, answer_image_html, ch_title, source_ref],
                tags=[f"chapter{int(ch_num):02d}"]
            )
            ch_deck.notes.append(note)

        # Add chapter deck to list if it has notes
        if ch_deck.notes:
            all_decks.append(ch_deck)

    # Add generated cloze cards if requested (organized by chapter)
    if include_generated and generated_questions:
        generated_cards = generated_questions.get("generated_cards", {})

        if generated_cards:
            # generated_cards is keyed by chapter (e.g., "ch4"), each value is a list of cards
            # Cards can have either:
            #   - source_question_id (legacy) - e.g., "ch4_2a"
            #   - source_block_id (block-based) - e.g., "ch4_2"

            # Build question lookup to get source explanations
            question_lookup = {}
            for ch_key_q, ch_qs in questions.items():
                for q in ch_qs:
                    question_lookup[q['full_id']] = q

            # Build block lookup - check both question_blocks and raw_blocks
            block_lookup = {}
            question_blocks = st.session_state.get("question_blocks", {})
            raw_blocks = st.session_state.get("raw_blocks", {})

            # Prefer raw_blocks as it has question_text_raw/answer_text_raw
            blocks_to_index = raw_blocks if raw_blocks else question_blocks
            if blocks_to_index:
                for ch_key_b, ch_blocks in blocks_to_index.items():
                    for block in ch_blocks:
                        block_id = block.get("block_id", block.get("block_label", ""))
                        if block_id:
                            block_lookup[block_id] = block

            # Create a generated sub-deck for each chapter
            for ch in chapters:
                ch_num = str(ch['chapter_number'])
                ch_key = f"ch{ch_num}"
                ch_title = ch.get('title', f'Chapter {ch_num}')

                ch_cards = generated_cards.get(ch_key, [])
                if not ch_cards:
                    continue

                # Skip if this chapter's generated cards are not selected
                # Default to True if no selection exists (backwards compatibility)
                if not export_selections.get(f"{ch_key}_generated", True):
                    continue

                # Create chapter::Generated sub-deck
                # Zero-pad chapter number for proper sorting
                gen_deck_name = f"{book_name}::{int(ch_num):02d}. {ch_title}::Generated"
                gen_deck_id = stable_id(gen_deck_name)
                gen_deck = genanki.Deck(gen_deck_id, gen_deck_name)

                for card in ch_cards:
                    cloze_text = card.get('cloze_text', '')
                    if not cloze_text:
                        continue

                    # Determine if block-based or question-based card
                    source_block_id = card.get('source_block_id', '')
                    source_q_id = card.get('source_question_id', '')

                    # Build source reference and extra info
                    extra_parts = []
                    source_ref = ""

                    if source_block_id:
                        # Block-based card
                        source_block = block_lookup.get(source_block_id, {})
                        local_id = source_block_id

                        if card.get('learning_point'):
                            extra_parts.append(f"<b>Learning point:</b> {card['learning_point']}")
                        if card.get('category'):
                            extra_parts.append(f"<b>Category:</b> {card['category']}")
                        if card.get('confidence'):
                            extra_parts.append(f"<b>Confidence:</b> {card['confidence']}")

                        # Build source reference with raw block text
                        source_ref_parts = [f'<span class="cloze-source-label">Generated from Block {source_block_id}</span>']

                        if source_block:
                            # Get raw text from the block (new format)
                            question_text_raw = source_block.get('question_text_raw', '')
                            answer_text_raw = source_block.get('answer_text_raw', '')

                            if question_text_raw or answer_text_raw:
                                source_ref_parts.append('<div class="cloze-source-text">')
                                if question_text_raw:
                                    # Escape HTML entities in raw text
                                    escaped_q = question_text_raw.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                                    source_ref_parts.append(f'<b>QUESTION:</b>\n{escaped_q}')
                                if answer_text_raw:
                                    escaped_a = answer_text_raw.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                                    if question_text_raw:
                                        source_ref_parts.append('\n\n')
                                    source_ref_parts.append(f'<b>ANSWER:</b>\n{escaped_a}')
                                source_ref_parts.append('</div>')
                            else:
                                # Fallback to old format fields
                                context_text = source_block.get('context_text', '')
                                shared_answer = source_block.get('shared_answer_text', '')
                                if context_text or shared_answer:
                                    source_ref_parts.append('<div class="cloze-source-text">')
                                    if context_text:
                                        escaped_ctx = context_text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                                        source_ref_parts.append(f'<b>CONTEXT:</b>\n{escaped_ctx}')
                                    if shared_answer:
                                        escaped_ans = shared_answer.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                                        if context_text:
                                            source_ref_parts.append('\n\n')
                                        source_ref_parts.append(f'<b>DISCUSSION:</b>\n{escaped_ans}')
                                    source_ref_parts.append('</div>')

                        source_ref = ''.join(source_ref_parts)
                    else:
                        # Legacy question-based card
                        local_id = source_q_id.split('_')[-1] if '_' in source_q_id else source_q_id

                        # Look up source question to get explanation
                        source_q = question_lookup.get(source_q_id, {})
                        source_explanation = source_q.get('explanation', '')
                        source_text = source_q.get('text', '')

                        if card.get('learning_point'):
                            extra_parts.append(f"<b>Learning point:</b> {card['learning_point']}")
                        if card.get('category'):
                            extra_parts.append(f"<b>Category:</b> {card['category']}")
                        if card.get('confidence'):
                            extra_parts.append(f"<b>Confidence:</b> {card['confidence']}")

                        # Build source reference with question/explanation text
                        source_ref_parts = [f'<span class="cloze-source-label">Generated from Q{local_id}</span>']

                        if source_text or source_explanation:
                            source_ref_parts.append('<div class="cloze-source-text">')
                            if source_text:
                                escaped_q = source_text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                                source_ref_parts.append(f'<b>QUESTION:</b>\n{escaped_q}')
                            if source_explanation:
                                escaped_exp = source_explanation.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                                if source_text:
                                    source_ref_parts.append('\n\n')
                                source_ref_parts.append(f'<b>EXPLANATION:</b>\n{escaped_exp}')
                            source_ref_parts.append('</div>')

                        source_ref = ''.join(source_ref_parts)

                    extra = '<br>'.join(extra_parts)

                    # Create cloze note
                    cloze_note = genanki.Note(
                        model=cloze_model,
                        fields=[cloze_text, extra, source_ref],
                        tags=[f"chapter{int(ch_num):02d}", "generated", card.get('category', 'general')]
                    )
                    gen_deck.notes.append(cloze_note)

                if gen_deck.notes:
                    all_decks.append(gen_deck)

    # Create package with all chapter decks
    output_dir = get_output_dir()
    safe_name = re.sub(r'[^\w\-]', '_', book_name)
    output_path = os.path.join(output_dir, f"{safe_name}.apkg")

    if not all_decks:
        raise ValueError("No cards to export")

    package = genanki.Package(all_decks)
    if media_files:
        package.media_files = media_files

    package.write_to_file(output_path)

    return output_path


# =============================================================================
# Step 6: Generate Questions
# =============================================================================

def render_generate_step():
    """Render the cloze card generation step."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from datetime import datetime
    from llm_extraction import generate_cloze_cards_llm, generate_cloze_cards_from_block_llm
    from state_management import save_generated_questions

    col1, col2 = st.columns([4, 1])
    with col1:
        st.header("Step 6: Generate Questions")
    with col2:
        if st.button("Reset To This Step", key="clear_generate", type="secondary",
                     help="Clear generated cloze cards"):
            clear_step_data("generate")
            st.success("Generated cards cleared")
            st.rerun()

    # Show cost from last completed step
    display_last_step_cost()

    # Check if block data is available (support both question_blocks and raw_blocks)
    question_blocks = st.session_state.get("question_blocks", {})
    raw_blocks = st.session_state.get("raw_blocks", {})
    has_blocks = (bool(question_blocks) and any(question_blocks.values())) or \
                 (bool(raw_blocks) and any(raw_blocks.values()))

    if has_blocks:
        st.markdown("""
        Generate cloze deletion flashcards from **question blocks**.
        Each block includes shared context and discussion, providing full context for accurate card generation.
        """)
        # Block-based generation toggle
        use_block_generation = st.checkbox(
            "Use block-based generation (recommended)",
            value=True,
            help="Generate cards from entire blocks with full context. "
                 "This prevents hallucination by providing the model with all related content."
        )
    else:
        use_block_generation = False
        st.markdown("""
        Generate cloze deletion flashcards from question explanations.
        Each explanation is analyzed to extract key learning points that become additional Anki cards.
        """)

    # Check prerequisites
    questions_source = st.session_state.questions
    if not questions_source:
        st.warning("Please format questions first (Step 4)")
        return

    client = get_anthropic_client()
    if not client:
        st.error("ANTHROPIC_API_KEY not set. Please set your API key in environment variables.")
        return

    # Gather blocks for block-based generation
    # Prefer raw_blocks (new format with question_text_raw/answer_text_raw) over question_blocks
    # Raw blocks can be fed directly to the LLM without additional formatting
    raw_blocks = st.session_state.get("raw_blocks", {})

    # Check if raw_blocks has the new format (answer_text_raw field)
    has_new_format_blocks = False
    if raw_blocks:
        for ch_blocks in raw_blocks.values():
            if ch_blocks and any(b.get("answer_text_raw") for b in ch_blocks):
                has_new_format_blocks = True
                break

    blocks_source = raw_blocks if has_new_format_blocks else (question_blocks if question_blocks else raw_blocks)
    all_blocks = []
    if use_block_generation and blocks_source:
        for ch_key in sort_chapter_keys(blocks_source.keys()):
            ch_blocks = blocks_source[ch_key]
            ch_num = int(ch_key[2:]) if ch_key.startswith("ch") else 0
            for block in ch_blocks:
                # Check if block has content worth generating from
                # Support both old format (shared_answer_text/sub_questions) and new format (answer_text_raw)
                has_content = (
                    block.get("answer_text_raw") or  # New block format
                    block.get("shared_answer_text") or  # Old format
                    any(sq.get("specific_answer_text") for sq in block.get("sub_questions", []))  # Old format
                )
                if has_content:
                    all_blocks.append((ch_key, ch_num, block))

    # Count questions with explanations (skip context-only) for legacy mode
    all_questions = []
    for ch_key in sort_chapter_keys(questions_source.keys()):
        ch_questions = questions_source[ch_key]
        ch_num = int(ch_key[2:]) if ch_key.startswith("ch") else 0
        for q in ch_questions:
            explanation = q.get("explanation", "")
            if explanation and len(explanation) >= 50 and not q.get("is_context_only"):
                all_questions.append((ch_key, ch_num, q))

    total_with_explanations = len(all_questions)
    total_blocks_to_process = len(all_blocks)
    generated_cards = st.session_state.generated_questions.get("generated_cards", {})
    total_generated = sum(len(cards) for cards in generated_cards.values())

    # Stats row
    col1, col2, col3 = st.columns(3)
    if use_block_generation:
        with col1:
            st.metric("Blocks to Process", total_blocks_to_process)
        with col2:
            st.metric("Generated Cards", total_generated)
        with col3:
            avg = total_generated / total_blocks_to_process if total_blocks_to_process > 0 else 0
            st.metric("Avg Cards/Block", f"{avg:.1f}")
    else:
        with col1:
            st.metric("Questions with Explanations", total_with_explanations)
        with col2:
            st.metric("Generated Cards", total_generated)
        with col3:
            avg = total_generated / total_with_explanations if total_with_explanations > 0 else 0
            st.metric("Avg Cards/Question", f"{avg:.1f}")

    st.markdown("---")

    # Chapter selector
    if not st.session_state.chapters:
        st.warning("No chapters available. Please extract chapters first (Step 2)")
        return

    chapter_options = [f"Ch{ch['chapter_number']}: {ch['title'][:40]}..."
                      for ch in st.session_state.chapters]
    selected_ch_idx = st.selectbox("Select chapter:",
                                    range(len(chapter_options)),
                                    format_func=lambda x: chapter_options[x],
                                    key="generate_ch_selector")

    selected_ch = st.session_state.chapters[selected_ch_idx]
    selected_ch_num = selected_ch["chapter_number"]
    selected_ch_key = f"ch{selected_ch_num}"

    # Count questions with explanations for selected chapter (legacy mode)
    ch_questions_with_explanations = [
        (ch_key, ch_num, q) for ch_key, ch_num, q in all_questions
        if ch_key == selected_ch_key
    ]
    ch_question_count = len(ch_questions_with_explanations)

    # Count blocks for selected chapter (block mode)
    ch_blocks_to_process = [
        (ch_key, ch_num, block) for ch_key, ch_num, block in all_blocks
        if ch_key == selected_ch_key
    ]
    ch_block_count = len(ch_blocks_to_process)

    # Generation controls
    model_col, workers_col, btn_col1, btn_col2 = st.columns([2, 1.5, 2, 2])

    with model_col:
        model_options = get_model_options()
        current_idx = model_options.index(st.session_state.selected_model) if st.session_state.selected_model in model_options else 0
        selected_model = st.selectbox("Model:", model_options, index=current_idx, key="generate_model")
        if selected_model != st.session_state.selected_model:
            st.session_state.selected_model = selected_model
            save_settings()

    with workers_col:
        gen_workers = st.number_input(
            "Parallel workers:",
            min_value=1, max_value=50, value=20,
            key="generate_workers",
            help="Number of questions to process in parallel"
        )

    with btn_col1:
        if use_block_generation:
            generate_chapter = st.button(
                f"Generate Chapter {selected_ch_num} ({ch_block_count} blocks)",
                type="primary",
                key="gen_chapter_btn",
                disabled=ch_block_count == 0
            )
        else:
            generate_chapter = st.button(
                f"Generate Chapter {selected_ch_num} ({ch_question_count})",
                type="primary",
                key="gen_chapter_btn",
                disabled=ch_question_count == 0
            )

    with btn_col2:
        if use_block_generation:
            generate_all = st.button(
                f"Generate ALL ({total_blocks_to_process} blocks)",
                type="secondary",
                key="gen_all_btn"
            )
        else:
            generate_all = st.button("Generate ALL Cards", type="secondary", key="gen_all_btn")

    # Single chapter generation logic
    if generate_chapter:
        model_id = get_selected_model_id()
        progress_bar = st.progress(0)
        status_text = st.empty()

        # Initialize generated_questions structure
        if not st.session_state.generated_questions.get("metadata"):
            st.session_state.generated_questions["metadata"] = {
                "created_at": datetime.now().isoformat(),
                "model_used": model_id,
                "total_generated": 0,
                "source_questions_processed": 0,
                "source_blocks_processed": 0
            }
        if "generated_cards" not in st.session_state.generated_questions:
            st.session_state.generated_questions["generated_cards"] = {}

        # Clear existing cards for this chapter
        if selected_ch_key in st.session_state.generated_questions["generated_cards"]:
            st.session_state.generated_questions["generated_cards"][selected_ch_key] = []

        completed = 0
        total_cards_generated = 0

        if use_block_generation:
            # Block-based generation
            if ch_block_count == 0:
                st.warning(f"No blocks to process in Chapter {selected_ch_num}")
            else:
                def process_block(item):
                    ch_key, ch_num, block = item
                    cards = generate_cloze_cards_from_block_llm(client, block, ch_num, model_id)
                    return ch_key, block, cards

                with ThreadPoolExecutor(max_workers=gen_workers) as executor:
                    futures = {executor.submit(process_block, item): item for item in ch_blocks_to_process}

                    for future in as_completed(futures):
                        ch_key, block, cards = future.result()
                        completed += 1

                        if cards:
                            # Initialize chapter list if needed
                            if ch_key not in st.session_state.generated_questions["generated_cards"]:
                                st.session_state.generated_questions["generated_cards"][ch_key] = []

                            # Build card objects with block source info
                            block_id = block.get("block_id", block.get("block_label", "unknown"))
                            for i, card in enumerate(cards, 1):
                                full_card = {
                                    "generated_id": f"{ch_key}_{block_id}_gen_{i}",
                                    "source_block_id": block_id,
                                    "cloze_text": card.get("cloze_text", ""),
                                    "learning_point": card.get("learning_point", ""),
                                    "confidence": card.get("confidence", "medium"),
                                    "category": card.get("category", "")
                                }
                                st.session_state.generated_questions["generated_cards"][ch_key].append(full_card)
                                total_cards_generated += 1

                        # Update progress
                        progress = completed / ch_block_count
                        progress_bar.progress(progress)
                        status_text.text(f"Chapter {selected_ch_num}: {completed}/{ch_block_count} blocks | {total_cards_generated} cards generated")

                        # Save periodically (every 5 blocks)
                        if completed % 5 == 0:
                            st.session_state.generated_questions["metadata"]["total_generated"] = sum(
                                len(cards) for cards in st.session_state.generated_questions["generated_cards"].values()
                            )
                            st.session_state.generated_questions["metadata"]["source_blocks_processed"] = completed
                            save_generated_questions()

                # Final save
                st.session_state.generated_questions["metadata"]["total_generated"] = sum(
                    len(cards) for cards in st.session_state.generated_questions["generated_cards"].values()
                )
                st.session_state.generated_questions["metadata"]["source_blocks_processed"] = completed
                save_generated_questions()

                save_cost_tracking(get_output_dir())
                progress_bar.progress(1.0)
                status_text.text(f"Complete: {completed} blocks processed, {total_cards_generated} cards generated")
                st.success(f"Generated {total_cards_generated} cloze cards from {completed} blocks in Chapter {selected_ch_num}!")
                render_cost_metrics("generate_cloze_block")
                save_cost_tracking(get_output_dir())
                st.rerun()
        else:
            # Legacy question-based generation
            if ch_question_count == 0:
                st.warning(f"No questions with explanations in Chapter {selected_ch_num}")
            else:
                def process_question(item):
                    ch_key, ch_num, q = item
                    cards = generate_cloze_cards_llm(client, q, ch_num, model_id)
                    return ch_key, q, cards

                with ThreadPoolExecutor(max_workers=gen_workers) as executor:
                    futures = {executor.submit(process_question, item): item for item in ch_questions_with_explanations}

                    for future in as_completed(futures):
                        ch_key, q, cards = future.result()
                        completed += 1

                        if cards:
                            # Initialize chapter list if needed
                            if ch_key not in st.session_state.generated_questions["generated_cards"]:
                                st.session_state.generated_questions["generated_cards"][ch_key] = []

                            # Build card objects
                            for i, card in enumerate(cards, 1):
                                full_card = {
                                    "generated_id": f"{q['full_id']}_gen_{i}",
                                    "source_question_id": q["full_id"],
                                    "cloze_text": card.get("cloze_text", ""),
                                    "learning_point": card.get("learning_point", ""),
                                    "confidence": card.get("confidence", "medium"),
                                    "category": card.get("category", "")
                                }
                                st.session_state.generated_questions["generated_cards"][ch_key].append(full_card)
                                total_cards_generated += 1

                        # Update progress
                        progress = completed / ch_question_count
                        progress_bar.progress(progress)
                        status_text.text(f"Chapter {selected_ch_num}: {completed}/{ch_question_count} questions | {total_cards_generated} cards generated")

                        # Save periodically (every 10 questions)
                        if completed % 10 == 0:
                            st.session_state.generated_questions["metadata"]["total_generated"] = sum(
                                len(cards) for cards in st.session_state.generated_questions["generated_cards"].values()
                            )
                            save_generated_questions()

                # Final save
                st.session_state.generated_questions["metadata"]["total_generated"] = sum(
                    len(cards) for cards in st.session_state.generated_questions["generated_cards"].values()
                )
                save_generated_questions()

                save_cost_tracking(get_output_dir())
                progress_bar.progress(1.0)
                status_text.text(f"Complete: {completed} questions processed, {total_cards_generated} cards generated")
                st.success(f"Generated {total_cards_generated} cloze cards from {completed} questions in Chapter {selected_ch_num}!")
                render_cost_metrics("generate_cloze")
                save_cost_tracking(get_output_dir())
                st.rerun()

    # All chapters generation logic
    if generate_all:
        model_id = get_selected_model_id()
        progress_bar = st.progress(0)
        status_text = st.empty()

        # Initialize generated_questions structure
        if not st.session_state.generated_questions.get("metadata"):
            st.session_state.generated_questions["metadata"] = {
                "created_at": datetime.now().isoformat(),
                "model_used": model_id,
                "total_generated": 0,
                "source_questions_processed": 0,
                "source_blocks_processed": 0
            }
        if "generated_cards" not in st.session_state.generated_questions:
            st.session_state.generated_questions["generated_cards"] = {}

        # Clear all existing cards
        st.session_state.generated_questions["generated_cards"] = {}

        completed = 0
        total_cards_generated = 0

        if use_block_generation:
            # Block-based generation for all chapters
            if total_blocks_to_process == 0:
                st.warning("No blocks found to process")
            else:
                def process_block(item):
                    ch_key, ch_num, block = item
                    cards = generate_cloze_cards_from_block_llm(client, block, ch_num, model_id)
                    return ch_key, block, cards

                with ThreadPoolExecutor(max_workers=gen_workers) as executor:
                    futures = {executor.submit(process_block, item): item for item in all_blocks}

                    for future in as_completed(futures):
                        ch_key, block, cards = future.result()
                        completed += 1

                        if cards:
                            # Initialize chapter list if needed
                            if ch_key not in st.session_state.generated_questions["generated_cards"]:
                                st.session_state.generated_questions["generated_cards"][ch_key] = []

                            # Build card objects with block source info
                            block_id = block.get("block_id", block.get("block_label", "unknown"))
                            for i, card in enumerate(cards, 1):
                                full_card = {
                                    "generated_id": f"{ch_key}_{block_id}_gen_{i}",
                                    "source_block_id": block_id,
                                    "cloze_text": card.get("cloze_text", ""),
                                    "learning_point": card.get("learning_point", ""),
                                    "confidence": card.get("confidence", "medium"),
                                    "category": card.get("category", "")
                                }
                                st.session_state.generated_questions["generated_cards"][ch_key].append(full_card)
                                total_cards_generated += 1

                        # Update progress
                        progress = completed / total_blocks_to_process
                        progress_bar.progress(progress)
                        status_text.text(f"Processing: {completed}/{total_blocks_to_process} blocks | {total_cards_generated} cards generated")

                        # Save periodically (every 5 blocks)
                        if completed % 5 == 0:
                            st.session_state.generated_questions["metadata"]["total_generated"] = total_cards_generated
                            st.session_state.generated_questions["metadata"]["source_blocks_processed"] = completed
                            save_generated_questions()

                # Final save
                st.session_state.generated_questions["metadata"]["total_generated"] = total_cards_generated
                st.session_state.generated_questions["metadata"]["source_blocks_processed"] = completed
                save_generated_questions()

                save_cost_tracking(get_output_dir())
                progress_bar.progress(1.0)
                status_text.text(f"Complete: {completed} blocks processed, {total_cards_generated} cards generated")
                st.success(f"Generated {total_cards_generated} cloze cards from {completed} blocks!")
                render_cost_metrics("generate_cloze_block")
                save_cost_tracking(get_output_dir())
                st.rerun()
        else:
            # Legacy question-based generation
            if total_with_explanations == 0:
                st.warning("No questions with explanations found to process")
            else:
                def process_question(item):
                    ch_key, ch_num, q = item
                    cards = generate_cloze_cards_llm(client, q, ch_num, model_id)
                    return ch_key, q, cards

                with ThreadPoolExecutor(max_workers=gen_workers) as executor:
                    futures = {executor.submit(process_question, item): item for item in all_questions}

                    for future in as_completed(futures):
                        ch_key, q, cards = future.result()
                        completed += 1

                        if cards:
                            # Initialize chapter list if needed
                            if ch_key not in st.session_state.generated_questions["generated_cards"]:
                                st.session_state.generated_questions["generated_cards"][ch_key] = []

                            # Build card objects
                            for i, card in enumerate(cards, 1):
                                full_card = {
                                    "generated_id": f"{q['full_id']}_gen_{i}",
                                    "source_question_id": q["full_id"],
                                    "cloze_text": card.get("cloze_text", ""),
                                    "learning_point": card.get("learning_point", ""),
                                    "confidence": card.get("confidence", "medium"),
                                    "category": card.get("category", "")
                                }
                                st.session_state.generated_questions["generated_cards"][ch_key].append(full_card)
                                total_cards_generated += 1

                        # Update progress
                        progress = completed / total_with_explanations
                        progress_bar.progress(progress)
                        status_text.text(f"Processing: {completed}/{total_with_explanations} questions | {total_cards_generated} cards generated")

                        # Save periodically (every 10 questions)
                        if completed % 10 == 0:
                            st.session_state.generated_questions["metadata"]["total_generated"] = total_cards_generated
                            st.session_state.generated_questions["metadata"]["source_questions_processed"] = completed
                            save_generated_questions()

                # Final save
                st.session_state.generated_questions["metadata"]["total_generated"] = total_cards_generated
                st.session_state.generated_questions["metadata"]["source_questions_processed"] = completed
                save_generated_questions()

                save_cost_tracking(get_output_dir())
                progress_bar.progress(1.0)
                status_text.text(f"Complete: {completed} questions processed, {total_cards_generated} cards generated")
                st.success(f"Generated {total_cards_generated} cloze cards from {completed} questions!")
                render_cost_metrics("generate_cloze")
                save_cost_tracking(get_output_dir())
                st.rerun()

    # Preview section
    generated_cards = st.session_state.generated_questions.get("generated_cards", {})
    if generated_cards:
        st.markdown("---")
        st.subheader("Generated Cards Preview")

        # Determine if using block-based cards
        first_ch_cards = next(iter(generated_cards.values()), [])
        is_block_based = first_ch_cards and "source_block_id" in first_ch_cards[0]

        # Filters
        col1, col2 = st.columns(2)
        with col1:
            ch_options = ["All Chapters"] + sort_chapter_keys(generated_cards.keys())
            ch_filter = st.selectbox("Filter by chapter:", ch_options, key="gen_ch_filter")
        with col2:
            # Get unique source IDs (block_id or question_id)
            source_ids = set()
            for ch_cards in generated_cards.values():
                for card in ch_cards:
                    if is_block_based:
                        source_ids.add(card.get("source_block_id", ""))
                    else:
                        source_ids.add(card.get("source_question_id", ""))
            filter_label = "Filter by block:" if is_block_based else "Filter by source:"
            all_label = "All Blocks" if is_block_based else "All Questions"
            source_options = [all_label] + sorted([s for s in source_ids if s], key=question_sort_key)
            source_filter = st.selectbox(filter_label, source_options, key="gen_source_filter")

        # Collect filtered cards
        filtered_cards = []
        for ch_key in sort_chapter_keys(generated_cards.keys()):
            if ch_filter != "All Chapters" and ch_key != ch_filter:
                continue
            for card in generated_cards[ch_key]:
                if is_block_based:
                    card_source = card.get("source_block_id", "")
                else:
                    card_source = card.get("source_question_id", "")
                all_label = "All Blocks" if is_block_based else "All Questions"
                if source_filter != all_label and card_source != source_filter:
                    continue
                filtered_cards.append((ch_key, card))

        if not filtered_cards:
            st.caption("No cards match filters")
        else:
            st.caption(f"Showing {len(filtered_cards)} cards")

            # Card navigation
            if "gen_preview_idx" not in st.session_state:
                st.session_state.gen_preview_idx = 0
            if st.session_state.gen_preview_idx >= len(filtered_cards):
                st.session_state.gen_preview_idx = 0

            current_idx = st.session_state.gen_preview_idx
            ch_key, card = filtered_cards[current_idx]

            # Navigation
            nav_col1, nav_col2, nav_col3 = st.columns([1, 3, 1])
            with nav_col1:
                if st.button("< Previous", disabled=(current_idx == 0), key="gen_prev"):
                    st.session_state.gen_preview_idx -= 1
                    st.rerun()
            with nav_col2:
                st.markdown(f"<div style='text-align: center;'>Card {current_idx + 1} of {len(filtered_cards)}</div>",
                           unsafe_allow_html=True)
            with nav_col3:
                if st.button("Next >", disabled=(current_idx >= len(filtered_cards) - 1), key="gen_next"):
                    st.session_state.gen_preview_idx += 1
                    st.rerun()

            # Side-by-side display
            left_col, right_col = st.columns(2)

            with left_col:
                st.markdown("#### Source")

                if is_block_based:
                    # Block-based card - show block info
                    source_block_id = card.get("source_block_id", "")
                    st.markdown(f"**Block:** `{source_block_id}`")

                    # Find the source block - check both question_blocks and raw_blocks
                    source_block = None
                    raw_blocks = st.session_state.get("raw_blocks", {})
                    blocks_to_search = [question_blocks, raw_blocks]

                    for blocks_source in blocks_to_search:
                        if source_block:
                            break
                        for ch_key, ch_blocks in blocks_source.items():
                            for block in ch_blocks:
                                if block.get("block_id") == source_block_id or block.get("block_label") == source_block_id:
                                    source_block = block
                                    break
                            if source_block:
                                break

                    if source_block:
                        # Build a combined view of block content
                        # Handle both new format (question_text_raw/answer_text_raw) and old format
                        source_text_parts = []

                        # New format: raw text fields
                        if source_block.get("question_text_raw"):
                            source_text_parts.append(f"QUESTION:\n{source_block['question_text_raw']}")
                        if source_block.get("answer_text_raw"):
                            source_text_parts.append(f"\nANSWER:\n{source_block['answer_text_raw']}")

                        # Old format fallback
                        if not source_text_parts:
                            if source_block.get("context_text"):
                                source_text_parts.append(f"CONTEXT:\n{source_block['context_text']}")
                            if source_block.get("sub_questions"):
                                for sq in source_block["sub_questions"]:
                                    sq_text = f"\nQ{sq.get('local_id', '')}: {sq.get('question_text', '')}"
                                    if sq.get("specific_answer_text"):
                                        sq_text += f"\nA: {sq['specific_answer_text']}"
                                    source_text_parts.append(sq_text)
                            if source_block.get("shared_answer_text"):
                                source_text_parts.append(f"\nSHARED DISCUSSION:\n{source_block['shared_answer_text']}")

                        source_text = "\n".join(source_text_parts) if source_text_parts else "(No content found)"
                        st.text_area("Block content:", source_text, height=300, disabled=True, key=f"src_block_{current_idx}")
                    else:
                        st.warning("Source block not found")
                else:
                    # Legacy question-based card
                    source_q_id = card.get("source_question_id", "")
                    st.markdown(f"**Question ID:** `{source_q_id}`")

                    source_q = None
                    # Parse chapter from question ID (e.g., "ch1_12b" -> "ch1")
                    if "_" in source_q_id:
                        source_ch_key = source_q_id.split("_")[0]
                        ch_questions = questions_source.get(source_ch_key, [])
                        for q in ch_questions:
                            if q.get("full_id") == source_q_id:
                                source_q = q
                                break

                    if source_q:
                        # Show full explanation text as the source reference
                        explanation = source_q.get("explanation", "")
                        if explanation:
                            st.text_area("Explanation (source):", explanation, height=300, disabled=True, key=f"src_exp_{current_idx}")
                        else:
                            st.caption("No explanation available")
                    else:
                        st.warning("Source question not found")

            with right_col:
                st.markdown("#### Generated Cloze Card")

                # Display cloze with visual highlighting
                cloze_text = card.get("cloze_text", "")
                import re
                # Convert {{c1::text::hint}} to **[text::hint]** for visual display (keep hint)
                display_text = re.sub(r'\{\{c\d+::([^}]+)\}\}', r'**[\1]**', cloze_text)
                # Convert <b>...</b> to **...** for Markdown bold rendering
                display_text = re.sub(r'<b>([^<]+)</b>', r'**\1**', display_text)
                st.markdown(f"**Cloze:**")
                st.markdown(display_text)

                st.markdown(f"**Learning Point:** {card.get('learning_point', 'N/A')}")

                col_a, col_b = st.columns(2)
                with col_a:
                    st.markdown(f"**Confidence:** {card.get('confidence', 'N/A')}")
                with col_b:
                    st.markdown(f"**Category:** {card.get('category', 'N/A')}")

                # Show all cards from same source
                if is_block_based:
                    current_source = card.get("source_block_id", "")
                    same_source_cards = [c for _, c in filtered_cards if c.get("source_block_id") == current_source]
                else:
                    current_source = card.get("source_question_id", "")
                    same_source_cards = [c for _, c in filtered_cards if c.get("source_question_id") == current_source]

                if len(same_source_cards) > 1:
                    expander_label = f"All {len(same_source_cards)} cards from this block" if is_block_based else f"All {len(same_source_cards)} cards from this source"
                    with st.expander(expander_label):
                        for i, sc in enumerate(same_source_cards, 1):
                            # Convert cloze syntax and <b> tags for display (keep hint)
                            display = re.sub(r'\{\{c\d+::([^}]+)\}\}', r'**[\1]**', sc.get("cloze_text", ""))
                            display = re.sub(r'<b>([^<]+)</b>', r'**\1**', display)
                            st.markdown(f"{i}. {display}")


# =============================================================================
# Step 7: Export
# =============================================================================

def render_export_step():
    """Render export step."""
    col1, col2 = st.columns([4, 1])
    with col1:
        st.header("Step 7: Export to Anki")
    with col2:
        if st.button("Clear Exports", key="clear_export", type="secondary", help="Delete exported .apkg files"):
            clear_step_data("export")
            st.success("Export files cleared")
            st.rerun()

    if not st.session_state.questions:
        st.warning("Please format questions first (Step 4)")
        return

    # Calculate stats
    total = sum(len(qs) for qs in st.session_state.questions.values())
    context_only = sum(1 for qs in st.session_state.questions.values()
                       for q in qs if q.get('is_context_only'))
    exportable = total - context_only

    reviewed = st.session_state.qc_progress.get("reviewed", {})
    approved = sum(1 for r in reviewed.values() if r.get("status") == "approved")

    # Count generated cards
    generated_cards = st.session_state.generated_questions.get("generated_cards", {})
    total_generated = sum(len(cards) for cards in generated_cards.values())

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Questions", total)
    with col2:
        st.metric("Exportable", exportable, help="Excludes context-only entries")
    with col3:
        st.metric("Approved (QC'd)", approved)
    with col4:
        st.metric("Generated Cloze", total_generated, help="Cloze cards from Step 7")

    st.markdown("---")

    # Book name setting
    default_name = Path(st.session_state.current_pdf).stem if st.session_state.current_pdf else "Textbook"
    default_name = default_name.replace('_', ' ').replace('-', ' ')

    book_name = st.text_input("Book Name (used as deck name)",
                              value=default_name,
                              help="This will be the parent deck name in Anki")

    # Export options
    col1, col2, col3 = st.columns(3)
    with col1:
        only_approved = st.checkbox("Only export approved questions",
                                   value=False,
                                   help="Only include questions marked as approved in QC")
    with col2:
        include_images = st.checkbox("Include images",
                                    value=True,
                                    help="Embed assigned images in cards")
    with col3:
        include_generated = st.checkbox("Include generated cloze cards",
                                        value=total_generated > 0,
                                        disabled=total_generated == 0,
                                        help="Add cloze cards from Step 7 as a sub-deck")

    st.markdown("---")

    # Build deck structure data
    deck_structure = []
    for ch in (st.session_state.chapters or []):
        ch_num = ch['chapter_number']
        ch_title = ch.get('title', f'Chapter {ch_num}')
        ch_key = f"ch{ch_num}"
        extracted_count = len([q for q in st.session_state.questions.get(ch_key, [])
                               if not q.get('is_context_only')])
        gen_count = len(generated_cards.get(ch_key, []))

        if extracted_count > 0 or gen_count > 0:
            deck_structure.append({
                "ch_num": ch_num,
                "ch_key": ch_key,
                "ch_title": ch_title,
                "extracted_count": extracted_count,
                "gen_count": gen_count
            })

    # Preview deck structure with checkboxes
    st.subheader("Select Content to Export")
    st.caption("Choose which chapters and card types to include in the export")

    if not deck_structure:
        st.warning("No content available to export")
    else:
        # Select/Deselect all buttons
        col1, col2, col3 = st.columns([1, 1, 4])
        with col1:
            if st.button("Select All", key="export_select_all"):
                for ch_data in deck_structure:
                    ch_key = ch_data["ch_key"]
                    if ch_data["extracted_count"] > 0:
                        st.session_state[f"export_{ch_key}_extracted"] = True
                    if ch_data["gen_count"] > 0:
                        st.session_state[f"export_{ch_key}_generated"] = True
                st.rerun()
        with col2:
            if st.button("Deselect All", key="export_deselect_all"):
                for ch_data in deck_structure:
                    ch_key = ch_data["ch_key"]
                    if ch_data["extracted_count"] > 0:
                        st.session_state[f"export_{ch_key}_extracted"] = False
                    if ch_data["gen_count"] > 0:
                        st.session_state[f"export_{ch_key}_generated"] = False
                st.rerun()

        st.markdown(f"**{book_name}**")

        for ch_data in deck_structure:
            ch_num = ch_data["ch_num"]
            ch_key = ch_data["ch_key"]
            ch_title = ch_data["ch_title"]
            extracted_count = ch_data["extracted_count"]
            gen_count = ch_data["gen_count"]

            # Initialize defaults if not set (default to True)
            extracted_key = f"export_{ch_key}_extracted"
            generated_key = f"export_{ch_key}_generated"
            if extracted_key not in st.session_state and extracted_count > 0:
                st.session_state[extracted_key] = True
            if generated_key not in st.session_state and gen_count > 0:
                st.session_state[generated_key] = True

            st.markdown(f"**Chapter {ch_num}: {ch_title}**")

            col1, col2 = st.columns(2)
            with col1:
                if extracted_count > 0:
                    st.checkbox(
                        f"Extracted ({extracted_count} cards)",
                        key=extracted_key
                    )
                else:
                    st.caption("No extracted cards")

            with col2:
                if gen_count > 0:
                    st.checkbox(
                        f"Generated ({gen_count} cards)",
                        key=generated_key,
                        disabled=not include_generated
                    )
                else:
                    st.caption("No generated cards")

        # Summary of what will be exported
        total_extracted_selected = sum(
            ch_data["extracted_count"] for ch_data in deck_structure
            if st.session_state.get(f"export_{ch_data['ch_key']}_extracted", False)
        )
        total_generated_selected = sum(
            ch_data["gen_count"] for ch_data in deck_structure
            if st.session_state.get(f"export_{ch_data['ch_key']}_generated", False) and include_generated
        )

        st.markdown("---")
        st.markdown(f"**Will export:** {total_extracted_selected} extracted + {total_generated_selected} generated = **{total_extracted_selected + total_generated_selected} total cards**")

    # Export button
    if st.button("Export to Anki Deck", type="primary"):
        if not book_name.strip():
            st.error("Please enter a book name")
            return

        with st.spinner("Generating Anki deck..."):
            try:
                questions_to_export = st.session_state.questions
                assignments_to_export = st.session_state.image_assignments

                # Build export_selections from checkbox keys in session state
                export_selections = {}
                for ch in (st.session_state.chapters or []):
                    ch_key = f"ch{ch['chapter_number']}"
                    export_selections[f"{ch_key}_extracted"] = st.session_state.get(f"export_{ch_key}_extracted", True)
                    export_selections[f"{ch_key}_generated"] = st.session_state.get(f"export_{ch_key}_generated", True)

                output_path = generate_anki_deck(
                    book_name=book_name.strip(),
                    questions=questions_to_export,
                    chapters=st.session_state.chapters or [],
                    image_assignments=assignments_to_export,
                    images=st.session_state.images,
                    include_images=include_images,
                    only_approved=only_approved,
                    qc_progress=st.session_state.qc_progress,
                    generated_questions=st.session_state.generated_questions,
                    include_generated=include_generated,
                    export_selections=export_selections
                )

                st.success(f"Deck exported successfully!")
                play_completion_sound()
                st.info(f"Saved to: `{output_path}`")

                # Provide download button
                with open(output_path, 'rb') as f:
                    st.download_button(
                        label="Download .apkg file",
                        data=f,
                        file_name=os.path.basename(output_path),
                        mime="application/octet-stream"
                    )

            except Exception as e:
                st.error(f"Export failed: {str(e)}")
                import traceback
                st.code(traceback.format_exc())


# =============================================================================
# Step 8: Edit Prompts
# =============================================================================

def render_prompts_step():
    """Render prompt editor step."""
    st.header("Step 8: Edit Prompts")

    st.caption("Edit the LLM prompts used for extraction. Changes are saved to `scripts/config/prompts.yaml`.")

    prompts = load_prompts()

    # Create tabs for each prompt
    prompt_names = list(prompts.keys())
    tabs = st.tabs([prompts[name].get("description", name) for name in prompt_names])

    for i, prompt_name in enumerate(prompt_names):
        with tabs[i]:
            prompt_data = prompts[prompt_name]
            description = prompt_data.get("description", "")
            current_prompt = prompt_data.get("prompt", "")

            st.markdown(f"**Prompt name:** `{prompt_name}`")
            st.caption(description)

            # Text area for editing
            new_prompt = st.text_area(
                "Prompt template",
                value=current_prompt,
                height=400,
                key=f"prompt_editor_{prompt_name}",
                help="Use {variable_name} syntax for variables that get filled in at runtime"
            )

            col1, col2 = st.columns([1, 4])
            with col1:
                if st.button("Save", key=f"save_{prompt_name}", type="primary"):
                    prompts[prompt_name]["prompt"] = new_prompt
                    save_prompts(prompts)
                    st.success(f"Saved {prompt_name}!")
            with col2:
                if st.button("Reset to saved", key=f"reset_{prompt_name}"):
                    reload_prompts()
                    st.rerun()

    st.markdown("---")
    st.subheader("Tips")
    st.markdown("""
    - **Variables**: Use `{variable_name}` syntax. Available variables depend on the prompt:
      - `identify_chapters`: `{page_index}`
      - `extract_qa_pairs`: `{chapter_num}`, `{chapter_text}`
      - `match_images_to_questions`: `{chapter_num}`, `{questions_text}`, `{images_text}`
      - `associate_context`: `{questions_summary}`
    - **JSON output**: Most prompts expect JSON output. Keep the output format instructions.
    - **Testing**: After editing, re-run the relevant extraction step to test your changes.
    """)


# =============================================================================
# Step 9: Stats
# =============================================================================

def scan_output_directories() -> list[dict]:
    """Scan output directory for all textbook folders with data."""
    output_base = Path("output")
    if not output_base.exists():
        return []

    textbooks = []
    for folder in sorted(output_base.iterdir()):
        if folder.is_dir() and not folder.name.startswith('.'):
            # Check if it has any relevant data files
            questions_file = folder / "questions_by_chapter.json"
            generated_file = folder / "generated_questions.json"

            if questions_file.exists() or generated_file.exists():
                textbooks.append({
                    "name": folder.name,
                    "path": folder,
                    "display_name": folder.name.replace('_', ' ')
                })

    return textbooks


def load_textbook_stats(textbook_path: Path) -> dict:
    """Load and compute statistics for a single textbook."""
    stats = {
        "extracted": {
            "total": 0,
            "by_chapter": {},
            "with_images": 0,
            "with_answer_images": 0,
            "multi_part": 0,
            "context_only": 0
        },
        "generated": {
            "total": 0,
            "by_chapter": {},
            "by_confidence": {"high": 0, "medium": 0, "low": 0},
            "by_category": {},
            "model_used": None,
            "last_updated": None,
            "blocks_processed": 0
        },
        "cost": {
            "total_cost": 0.0,
            "total_input_tokens": 0,
            "total_output_tokens": 0,
            "by_step": {}
        }
    }

    # Load extracted questions
    questions_file = textbook_path / "questions_by_chapter.json"
    if questions_file.exists():
        try:
            with open(questions_file, 'r') as f:
                questions_by_chapter = json.load(f)

            for chapter, questions in questions_by_chapter.items():
                chapter_count = len(questions)
                stats["extracted"]["by_chapter"][chapter] = {
                    "count": chapter_count,
                    "with_images": 0,
                    "with_answer_images": 0,
                    "multi_part": 0,
                    "context_only": 0
                }
                stats["extracted"]["total"] += chapter_count

                for q in questions:
                    if q.get("image_files"):
                        stats["extracted"]["with_images"] += 1
                        stats["extracted"]["by_chapter"][chapter]["with_images"] += 1
                    if q.get("answer_image_files"):
                        stats["extracted"]["with_answer_images"] += 1
                        stats["extracted"]["by_chapter"][chapter]["with_answer_images"] += 1
                    if q.get("context_from"):
                        stats["extracted"]["multi_part"] += 1
                        stats["extracted"]["by_chapter"][chapter]["multi_part"] += 1
                    if q.get("is_context_only"):
                        stats["extracted"]["context_only"] += 1
                        stats["extracted"]["by_chapter"][chapter]["context_only"] += 1
        except (json.JSONDecodeError, IOError):
            pass

    # Load generated questions
    generated_file = textbook_path / "generated_questions.json"
    if generated_file.exists():
        try:
            with open(generated_file, 'r') as f:
                generated_data = json.load(f)

            metadata = generated_data.get("metadata", {})
            stats["generated"]["model_used"] = metadata.get("model_used")
            stats["generated"]["last_updated"] = metadata.get("last_updated") or metadata.get("created_at")
            stats["generated"]["blocks_processed"] = metadata.get("source_blocks_processed", 0)

            generated_cards = generated_data.get("generated_cards", {})
            for chapter, cards in generated_cards.items():
                chapter_count = len(cards)
                stats["generated"]["by_chapter"][chapter] = {
                    "count": chapter_count,
                    "high": 0,
                    "medium": 0,
                    "low": 0
                }
                stats["generated"]["total"] += chapter_count

                for card in cards:
                    confidence = card.get("confidence", "").lower()
                    if confidence in ["high", "medium", "low"]:
                        stats["generated"]["by_confidence"][confidence] += 1
                        stats["generated"]["by_chapter"][chapter][confidence] += 1

                    category = card.get("category", "uncategorized")
                    stats["generated"]["by_category"][category] = stats["generated"]["by_category"].get(category, 0) + 1
        except (json.JSONDecodeError, IOError):
            pass

    # Load cost tracking data
    cost_file = textbook_path / "cost_tracking.json"
    if cost_file.exists():
        try:
            with open(cost_file, 'r') as f:
                cost_data = json.load(f)

            stats["cost"]["total_cost"] = cost_data.get("total_cost", 0.0)
            stats["cost"]["total_input_tokens"] = cost_data.get("total_input_tokens", 0)
            stats["cost"]["total_output_tokens"] = cost_data.get("total_output_tokens", 0)
            stats["cost"]["by_step"] = cost_data.get("by_step", {})
        except (json.JSONDecodeError, IOError):
            pass

    return stats


def generate_excel_report(all_stats: list[tuple[str, dict]]) -> bytes:
    """Generate Excel report with summary and per-textbook sheets."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Summary headers
    summary_headers = ["Textbook", "Extracted", "Generated", "Ratio", "API Cost"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Summary data
    total_extracted = 0
    total_generated = 0
    total_cost = 0.0
    for row, (name, stats) in enumerate(all_stats, 2):
        extracted = stats["extracted"]["total"]
        generated = stats["generated"]["total"]
        cost = stats["cost"]["total_cost"]
        ratio = f"{(generated / extracted * 100):.0f}%" if extracted > 0 else "N/A"

        total_extracted += extracted
        total_generated += generated
        total_cost += cost

        display_name = name.replace('_', ' ')
        ws_summary.cell(row=row, column=1, value=display_name).border = thin_border
        ws_summary.cell(row=row, column=2, value=extracted).border = thin_border
        ws_summary.cell(row=row, column=3, value=generated).border = thin_border
        ws_summary.cell(row=row, column=4, value=ratio).border = thin_border
        ws_summary.cell(row=row, column=5, value=f"${cost:.2f}").border = thin_border

    # Total row
    total_row = len(all_stats) + 2
    total_ratio = f"{(total_generated / total_extracted * 100):.0f}%" if total_extracted > 0 else "N/A"
    for col, value in enumerate(["TOTAL", total_extracted, total_generated, total_ratio, f"${total_cost:.2f}"], 1):
        cell = ws_summary.cell(row=total_row, column=col, value=value)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border

    # Adjust column widths for summary
    ws_summary.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D', 'E']:
        ws_summary.column_dimensions[col].width = 12

    # Per-textbook sheets
    for name, stats in all_stats:
        # Sanitize sheet name (max 31 chars, no special chars)
        sheet_name = name.replace('_', ' ')[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Extracted Questions section
        ws.cell(row=1, column=1, value="Extracted Questions").font = Font(bold=True, size=12)
        ws.merge_cells('A1:E1')

        extracted_headers = ["Chapter", "Count", "With Images", "Multi-part", "Context-only"]
        for col, header in enumerate(extracted_headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Sort chapters naturally
        extracted_chapters = sorted(
            stats["extracted"]["by_chapter"].items(),
            key=lambda x: (int(x[0].replace('ch', '')) if x[0].replace('ch', '').isdigit() else 999, x[0])
        )

        row = 3
        ext_totals = {"count": 0, "with_images": 0, "multi_part": 0, "context_only": 0}
        for chapter, ch_stats in extracted_chapters:
            ws.cell(row=row, column=1, value=chapter).border = thin_border
            ws.cell(row=row, column=2, value=ch_stats["count"]).border = thin_border

            img_pct = f"{ch_stats['with_images']} ({ch_stats['with_images']/ch_stats['count']*100:.0f}%)" if ch_stats['count'] > 0 else "0"
            ws.cell(row=row, column=3, value=img_pct).border = thin_border
            ws.cell(row=row, column=4, value=ch_stats["multi_part"]).border = thin_border
            ws.cell(row=row, column=5, value=ch_stats["context_only"]).border = thin_border

            ext_totals["count"] += ch_stats["count"]
            ext_totals["with_images"] += ch_stats["with_images"]
            ext_totals["multi_part"] += ch_stats["multi_part"]
            ext_totals["context_only"] += ch_stats["context_only"]
            row += 1

        # Extracted totals
        total_img_pct = f"{ext_totals['with_images']} ({ext_totals['with_images']/ext_totals['count']*100:.0f}%)" if ext_totals['count'] > 0 else "0"
        for col, value in enumerate(["TOTAL", ext_totals["count"], total_img_pct, ext_totals["multi_part"], ext_totals["context_only"]], 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border

        # Generated Questions section (start a few rows down)
        gen_start_row = row + 3
        ws.cell(row=gen_start_row, column=1, value="Generated Questions").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{gen_start_row}:E{gen_start_row}')

        gen_headers = ["Chapter", "Count", "High", "Medium", "Low"]
        for col, header in enumerate(gen_headers, 1):
            cell = ws.cell(row=gen_start_row + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Sort generated chapters naturally
        generated_chapters = sorted(
            stats["generated"]["by_chapter"].items(),
            key=lambda x: (int(x[0].replace('ch', '')) if x[0].replace('ch', '').isdigit() else 999, x[0])
        )

        row = gen_start_row + 2
        gen_totals = {"count": 0, "high": 0, "medium": 0, "low": 0}
        for chapter, ch_stats in generated_chapters:
            ws.cell(row=row, column=1, value=chapter).border = thin_border
            ws.cell(row=row, column=2, value=ch_stats["count"]).border = thin_border
            ws.cell(row=row, column=3, value=ch_stats["high"]).border = thin_border
            ws.cell(row=row, column=4, value=ch_stats["medium"]).border = thin_border
            ws.cell(row=row, column=5, value=ch_stats["low"]).border = thin_border

            gen_totals["count"] += ch_stats["count"]
            gen_totals["high"] += ch_stats["high"]
            gen_totals["medium"] += ch_stats["medium"]
            gen_totals["low"] += ch_stats["low"]
            row += 1

        # Generated totals
        for col, value in enumerate(["TOTAL", gen_totals["count"], gen_totals["high"], gen_totals["medium"], gen_totals["low"]], 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border

        # Metadata section
        meta_row = row + 3
        if stats["generated"]["model_used"]:
            ws.cell(row=meta_row, column=1, value="Model:").font = Font(bold=True)
            ws.cell(row=meta_row, column=2, value=stats["generated"]["model_used"])
            meta_row += 1
        if stats["generated"]["last_updated"]:
            ws.cell(row=meta_row, column=1, value="Last Updated:").font = Font(bold=True)
            ws.cell(row=meta_row, column=2, value=stats["generated"]["last_updated"])
            meta_row += 1

        # Cost Breakdown section
        if stats["cost"]["total_cost"] > 0:
            cost_start_row = meta_row + 2
            ws.cell(row=cost_start_row, column=1, value="API Cost Breakdown").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{cost_start_row}:E{cost_start_row}')

            cost_headers = ["Step", "Input Tokens", "Output Tokens", "Calls", "Cost"]
            for col, header in enumerate(cost_headers, 1):
                cell = ws.cell(row=cost_start_row + 1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            cost_row = cost_start_row + 2
            cost_totals = {"input": 0, "output": 0, "calls": 0, "cost": 0.0}
            for step_name, step_stats in stats["cost"]["by_step"].items():
                ws.cell(row=cost_row, column=1, value=step_name.replace('_', ' ').title()).border = thin_border
                ws.cell(row=cost_row, column=2, value=step_stats["input"]).border = thin_border
                ws.cell(row=cost_row, column=3, value=step_stats["output"]).border = thin_border
                ws.cell(row=cost_row, column=4, value=step_stats.get("call_count", 1)).border = thin_border
                ws.cell(row=cost_row, column=5, value=f"${step_stats['cost']:.4f}").border = thin_border

                cost_totals["input"] += step_stats["input"]
                cost_totals["output"] += step_stats["output"]
                cost_totals["calls"] += step_stats.get("call_count", 1)
                cost_totals["cost"] += step_stats["cost"]
                cost_row += 1

            # Cost totals
            for col, value in enumerate(["TOTAL", cost_totals["input"], cost_totals["output"], cost_totals["calls"], f"${cost_totals['cost']:.4f}"], 1):
                cell = ws.cell(row=cost_row, column=col, value=value)
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border

        # Adjust column widths
        ws.column_dimensions['A'].width = 18
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 14

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def render_stats_step():
    """Render statistics overview for all textbooks."""
    st.header("Step 9: Stats")
    st.caption("View extraction and generation statistics across all processed textbooks.")

    # Scan for textbooks
    textbooks = scan_output_directories()

    if not textbooks:
        st.warning("No textbooks with data found in the output directory.")
        st.info("Process at least one textbook through Steps 1-4 to see statistics here.")
        return

    # Load stats for all textbooks
    all_stats = []
    for tb in textbooks:
        stats = load_textbook_stats(tb["path"])
        all_stats.append((tb["name"], stats))

    # Calculate totals
    total_extracted = sum(s["extracted"]["total"] for _, s in all_stats)
    total_generated = sum(s["generated"]["total"] for _, s in all_stats)
    total_cost = sum(s["cost"]["total_cost"] for _, s in all_stats)

    # Excel export button at top
    col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
    with col1:
        excel_data = generate_excel_report(all_stats)
        st.download_button(
            label="Download Excel Report",
            data=excel_data,
            file_name=f"textbook_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
    with col2:
        st.metric("Total Extracted", f"{total_extracted:,}")
    with col3:
        st.metric("Total Generated", f"{total_generated:,}")
    with col4:
        st.metric("Total API Cost", f"${total_cost:.2f}")

    st.markdown("---")

    # Create tabs: Summary + one per textbook
    tab_names = ["Summary"] + [tb["display_name"] for tb in textbooks]
    tabs = st.tabs(tab_names)

    # Summary tab
    with tabs[0]:
        st.subheader("All Textbooks Summary")

        # Build summary dataframe
        summary_data = []
        for name, stats in all_stats:
            extracted = stats["extracted"]["total"]
            generated = stats["generated"]["total"]
            cost = stats["cost"]["total_cost"]
            ratio = f"{(generated / extracted * 100):.0f}%" if extracted > 0 else "N/A"
            summary_data.append({
                "Textbook": name.replace('_', ' '),
                "Extracted": extracted,
                "Generated": generated,
                "Ratio": ratio,
                "API Cost": f"${cost:.2f}"
            })

        # Add total row
        total_ratio = f"{(total_generated / total_extracted * 100):.0f}%" if total_extracted > 0 else "N/A"
        summary_data.append({
            "Textbook": "TOTAL",
            "Extracted": total_extracted,
            "Generated": total_generated,
            "Ratio": total_ratio,
            "API Cost": f"${total_cost:.2f}"
        })

        st.dataframe(
            summary_data,
            use_container_width=True,
            hide_index=True
        )

    # Per-textbook tabs
    for i, (tb, (name, stats)) in enumerate(zip(textbooks, all_stats)):
        with tabs[i + 1]:
            st.subheader(f"{tb['display_name']} Details")

            # Two columns: Extracted and Generated
            col1, col2 = st.columns(2)

            with col1:
                st.markdown("**Extracted Questions**")

                if stats["extracted"]["total"] > 0:
                    # Build extracted data
                    extracted_data = []
                    sorted_chapters = sorted(
                        stats["extracted"]["by_chapter"].items(),
                        key=lambda x: (int(x[0].replace('ch', '')) if x[0].replace('ch', '').isdigit() else 999, x[0])
                    )

                    for chapter, ch_stats in sorted_chapters:
                        count = ch_stats["count"]
                        img_pct = f"{ch_stats['with_images']} ({ch_stats['with_images']/count*100:.0f}%)" if count > 0 else "0"
                        extracted_data.append({
                            "Chapter": chapter,
                            "Count": count,
                            "With Images": img_pct,
                            "Multi-part": ch_stats["multi_part"],
                            "Context-only": ch_stats["context_only"]
                        })

                    # Add totals
                    total_count = stats["extracted"]["total"]
                    total_with_img = stats["extracted"]["with_images"]
                    total_img_pct = f"{total_with_img} ({total_with_img/total_count*100:.0f}%)" if total_count > 0 else "0"
                    extracted_data.append({
                        "Chapter": "TOTAL",
                        "Count": total_count,
                        "With Images": total_img_pct,
                        "Multi-part": stats["extracted"]["multi_part"],
                        "Context-only": stats["extracted"]["context_only"]
                    })

                    st.dataframe(extracted_data, use_container_width=True, hide_index=True)
                else:
                    st.info("No extracted questions.")

            with col2:
                st.markdown("**Generated Questions**")

                if stats["generated"]["total"] > 0:
                    # Build generated data
                    generated_data = []
                    sorted_chapters = sorted(
                        stats["generated"]["by_chapter"].items(),
                        key=lambda x: (int(x[0].replace('ch', '')) if x[0].replace('ch', '').isdigit() else 999, x[0])
                    )

                    for chapter, ch_stats in sorted_chapters:
                        generated_data.append({
                            "Chapter": chapter,
                            "Count": ch_stats["count"],
                            "High": ch_stats["high"],
                            "Medium": ch_stats["medium"],
                            "Low": ch_stats["low"]
                        })

                    # Add totals
                    conf = stats["generated"]["by_confidence"]
                    generated_data.append({
                        "Chapter": "TOTAL",
                        "Count": stats["generated"]["total"],
                        "High": conf["high"],
                        "Medium": conf["medium"],
                        "Low": conf["low"]
                    })

                    st.dataframe(generated_data, use_container_width=True, hide_index=True)

                    # Metadata
                    if stats["generated"]["model_used"] or stats["generated"]["last_updated"]:
                        st.markdown("---")
                        meta_cols = st.columns(2)
                        if stats["generated"]["model_used"]:
                            with meta_cols[0]:
                                st.caption(f"**Model:** {stats['generated']['model_used']}")
                        if stats["generated"]["last_updated"]:
                            with meta_cols[1]:
                                st.caption(f"**Last Updated:** {stats['generated']['last_updated']}")
                else:
                    st.info("No generated questions.")

            # Cost breakdown section
            if stats["cost"]["total_cost"] > 0:
                st.markdown("---")
                st.markdown("**API Cost Breakdown**")

                cost_data = []
                for step_name, step_stats in stats["cost"]["by_step"].items():
                    cost_data.append({
                        "Step": step_name.replace('_', ' ').title(),
                        "Input Tokens": f"{step_stats['input']:,}",
                        "Output Tokens": f"{step_stats['output']:,}",
                        "Calls": step_stats.get("call_count", 1),
                        "Cost": f"${step_stats['cost']:.4f}"
                    })

                # Add total row
                cost_data.append({
                    "Step": "TOTAL",
                    "Input Tokens": f"{stats['cost']['total_input_tokens']:,}",
                    "Output Tokens": f"{stats['cost']['total_output_tokens']:,}",
                    "Calls": sum(s.get("call_count", 1) for s in stats["cost"]["by_step"].values()),
                    "Cost": f"${stats['cost']['total_cost']:.4f}"
                })

                st.dataframe(cost_data, use_container_width=True, hide_index=True)
